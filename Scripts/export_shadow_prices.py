#!/usr/bin/env python3
"""
export_shadow_prices.py

Export the shadow price of power from the results DB, converted to EUR/MWh.

Source: duals_EB2_EnergyBalanceEachTS - the dual of the energy-balance
constraint (= marginal / shadow price of the commodity node).  The dimensions
are pipe-encoded in the Constraint string:

    EB2_EnergyBalanceEachTS | Year | Timeslice | Fuel | Region

and Dual is the shadow price in M EUR / PJ.

Unit conversion:  1 M EUR/PJ = 1e6 EUR / (1e15 J / 3.6e9 J per MWh)
                             = 1e6 / 277 777.8 = 3.6 EUR/MWh.

By default only Fuel='Power' (the wholesale electricity node) is exported;
the sector sub-nodes (Power_DataCenter, Power_Hydrogen, ...) can be added
with --all-fuels or picked with --fuel.

Output sheets:
    Shadow Prices (TS)   Scenario | Region | Year | Timeslice | Fuel | Price_EUR_MWh
    Annual Average       Scenario | Region | Year | Fuel | AvgPrice_EUR_MWh
                         (simple mean over timeslices - NOT load-weighted)

Run:
    python export_shadow_prices.py
    python export_shadow_prices.py --scenarios all
    python export_shadow_prices.py --all-fuels
    python export_shadow_prices.py --fuel Power_DataCenter -o dc_prices.xlsx
"""

import os
import sys
import argparse
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(SCRIPT_DIR, "..", "Results")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "Output")
RESULTS_DB = os.path.join(RESULTS_DIR, "genesysmod_db.duckdb")

DUALS_TABLE = "duals_EB2_EnergyBalanceEachTS"
GEN_TABLE = "varpar_Production"   # timeslice-resolved production (weights)
MEUR_PJ_TO_EUR_MWH = 3.6          # 1 M EUR/PJ = 3.6 EUR/MWh
POWER_FUEL = "Power"

US_REGIONS = ["California", "ERCOT", "MISO", "NewEngland", "NewYork",
              "PJM", "SERC", "SPP", "WECC"]
CANADA = "Canada"
AGGREGATES = {                    # generation-weighted groups
    "Total US": set(US_REGIONS),
    "Total North America": set(US_REGIONS) | {CANADA},
}

# region display order (US regions then Canada)
REGION_ORDER = {r: i for i, r in enumerate(
    ["California", "ERCOT", "MISO", "NewEngland", "NewYork",
     "PJM", "SERC", "SPP", "WECC", "Canada"])}

# --------------------------------------------------------------------------- #
# Scenario selection (multi)
# --------------------------------------------------------------------------- #
def distinct_scenarios(con):
    return sorted({(s, p, ps) for s, p, ps in con.execute(
        f"SELECT DISTINCT Scenario, Pathway, PathwayScenario FROM {DUALS_TABLE}"
    ).fetchall()})


def scenario_clause(combo):
    s, p, ps = combo
    esc = lambda x: str(x).replace("'", "''")
    return (f"Scenario='{esc(s)}' AND Pathway='{esc(p)}' "
            f"AND PathwayScenario='{esc(ps)}'")


def _match(scenarios, token):
    token = token.strip()
    if token.isdigit() and 1 <= int(token) <= len(scenarios):
        return scenarios[int(token) - 1]
    for c in scenarios:
        if token == c[0] or token == c[2]:
            return c
    return None


def choose_scenarios(scenarios, preset=None):
    if not scenarios:
        sys.exit("[error] no scenarios in duals table.")
    if preset:
        if preset.strip().lower() == "all":
            return scenarios
        chosen = [_match(scenarios, t) for t in preset.split(",")]
        if any(c is None for c in chosen):
            sys.exit(f"[error] unknown scenario in --scenarios '{preset}'.")
        return chosen
    if len(scenarios) == 1:
        print(f"  single scenario -> {scenarios[0][0]}")
        return scenarios
    print("\nAvailable scenarios:")
    for i, (s, p, ps) in enumerate(scenarios, 1):
        print(f"  [{i}] Scenario={s} | PathwayScenario={ps}")
    while True:
        try:
            raw = input("Select scenarios (comma-separated, or 'all'): ").strip()
        except EOFError:
            sys.exit("[error] no input. Pass --scenarios.")
        if raw.lower() == "all":
            return scenarios
        chosen = [_match(scenarios, t) for t in raw.split(",") if t.strip()]
        if chosen and all(c is not None for c in chosen):
            seen, out = set(), []
            for c in chosen:
                if c not in seen:
                    seen.add(c); out.append(c)
            return out
        print("  invalid selection, try again.")


# --------------------------------------------------------------------------- #
# Query shadow prices (parsed + converted)
# --------------------------------------------------------------------------- #
def fetch_prices(con, where, fuel_filter):
    """Return list of (Scenario, Region, Year, Timeslice, Fuel, price_EUR_MWh)."""
    q = f"""
        SELECT
            Scenario,
            CAST(split_part("Constraint", '|', 2) AS INTEGER) AS Year,
            split_part("Constraint", '|', 3)                  AS Timeslice,
            split_part("Constraint", '|', 4)                  AS Fuel,
            split_part("Constraint", '|', 5)                  AS Region,
            Dual * {MEUR_PJ_TO_EUR_MWH}                       AS Price
        FROM {DUALS_TABLE}
        WHERE {where}
    """
    rows = con.execute(q).fetchall()
    out = []
    for scen, year, ts, fuel, region, price in rows:
        if fuel_filter is not None and fuel not in fuel_filter:
            continue
        out.append((scen, region, year, ts, fuel, price))
    return out


def fetch_generation(con, where, fuel_filter):
    """Timeslice generation weights keyed (scen, region, year, ts_str, fuel)."""
    q = (f"SELECT Scenario, Region, Year, Timeslice, Fuel, Value "
         f"FROM {GEN_TABLE} WHERE {where}")
    gen = {}
    for scen, region, year, ts, fuel, val in con.execute(q).fetchall():
        if fuel_filter is not None and fuel not in fuel_filter:
            continue
        if val is None:
            continue
        gen[(scen, region, int(year), str(ts), fuel)] = \
            gen.get((scen, region, int(year), str(ts), fuel), 0.0) + val
    return gen


# --------------------------------------------------------------------------- #
# Workbook
# --------------------------------------------------------------------------- #
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")


def _sheet(ws, headers, rows, num_cols):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(r)
    for c in num_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    for c in range(1, len(headers) + 1):
        w = max(len(str(headers[c - 1])) + 2,
                max((len(str(ws.cell(r, c).value))
                     for r in range(2, min(ws.max_row, 200) + 1)), default=8) + 2)
        ws.column_dimensions[get_column_letter(c)].width = min(w, 40)


def build_workbook(prices, gen, out_path):
    wb = Workbook()
    wb.remove(wb.active)

    def sortkey(r):  # Scenario, Region order, Year, Fuel, Timeslice
        return (r[0], REGION_ORDER.get(r[1], 99), r[2], r[4], r[3])
    ts_rows = sorted(prices, key=sortkey)
    ts_out = [(s, reg, y, ts, fuel, round(p, 4))
              for (s, reg, y, ts, fuel, p) in ts_rows]

    # per-region annual average (simple mean over timeslices)
    acc = {}   # (scen, region, year, fuel) -> [sum, n]
    for s, reg, y, ts, fuel, p in prices:
        k = (s, reg, y, fuel)
        a = acc.setdefault(k, [0.0, 0])
        a[0] += p; a[1] += 1
    ann = [(s, reg, y, fuel, round(v[0] / v[1], 4))
           for (s, reg, y, fuel), v in acc.items()]
    ann.sort(key=lambda r: (r[0], REGION_ORDER.get(r[1], 99), r[2], r[3]))

    # generation-weighted average for Total US / Total North America
    # P = sum(price_ts,region * gen_ts,region) / sum(gen_ts,region)
    wacc = {}  # (scen, aggname, year, fuel) -> [wsum, gsum]
    for s, reg, y, ts, fuel, p in prices:
        g = gen.get((s, reg, y, str(ts), fuel))
        if g is None:
            continue
        for aggname, members in AGGREGATES.items():
            if reg in members:
                a = wacc.setdefault((s, aggname, y, fuel), [0.0, 0.0])
                a[0] += p * g
                a[1] += g
    agg_order = {"Total US": 0, "Total North America": 1}
    wtd = [(s, agg, y, fuel, round(w[0] / w[1], 4))
           for (s, agg, y, fuel), w in wacc.items() if w[1] > 0]
    wtd.sort(key=lambda r: (r[0], agg_order.get(r[1], 9), r[2], r[3]))

    _sheet(wb.create_sheet("Shadow Prices (TS)"),
           ["Scenario", "Region", "Year", "Timeslice", "Fuel", "Price_EUR_MWh"],
           ts_out, num_cols=[6])
    _sheet(wb.create_sheet("Gen-Weighted US & NA"),
           ["Scenario", "Aggregate", "Year", "Fuel", "GenWtdPrice_EUR_MWh"],
           wtd, num_cols=[5])
    _sheet(wb.create_sheet("Annual Average"),
           ["Scenario", "Region", "Year", "Fuel", "AvgPrice_EUR_MWh"],
           ann, num_cols=[5])

    # notes
    wsw = wb["Gen-Weighted US & NA"]
    n1 = wsw.cell(wsw.max_row + 2, 1)
    n1.value = ("Generation-weighted: sum(price_ts,region * generation_ts,region)"
                " / sum(generation_ts,region), over all timeslices & member "
                "regions (weights = varpar_Production, Fuel node).")
    n1.font = Font(italic=True, size=9, color="808080")

    ws = wb["Annual Average"]
    note = ws.cell(ws.max_row + 2, 1)
    note.value = ("Note: simple arithmetic mean over timeslices "
                  "(NOT load- or duration-weighted). See 'Gen-Weighted US & NA' "
                  "for generation-weighted aggregates.")
    note.font = Font(italic=True, size=9, color="808080")

    wb.save(out_path)
    return len(ts_out), len(ann), len(wtd)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(
        description="Export power shadow prices (M EUR/PJ -> EUR/MWh).")
    ap.add_argument("--scenarios", default=os.environ.get("GMOD_SCENARIOS"),
                    help="'all' or comma-separated Scenario/PathwayScenario names.")
    ap.add_argument("--fuel", default=None,
                    help=f"Commodity to export (default '{POWER_FUEL}').")
    ap.add_argument("--all-fuels", action="store_true",
                    help="Export every Power* commodity (Fuel column distinguishes).")
    ap.add_argument("-o", "--output", default=None)
    args = ap.parse_args()

    if not os.path.exists(RESULTS_DB):
        sys.exit(f"Missing database: {RESULTS_DB}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    con = duckdb.connect(RESULTS_DB, read_only=True)
    chosen = choose_scenarios(distinct_scenarios(con), preset=args.scenarios)
    where = "(" + " OR ".join(f"({scenario_clause(c)})" for c in chosen) + ")"

    if args.all_fuels:
        fuel_filter = None
    elif args.fuel:
        fuel_filter = {args.fuel}
    else:
        fuel_filter = {POWER_FUEL}

    prices = fetch_prices(con, where, fuel_filter)
    gen = fetch_generation(con, where, fuel_filter)
    con.close()
    if not prices:
        sys.exit("[error] no rows for the chosen scenarios/fuel.")

    out_path = args.output or os.path.join(OUTPUT_DIR, "Shadow_Prices_Power.xlsx")
    n_ts, n_ann, n_wtd = build_workbook(prices, gen, out_path)

    print(f"\n  scenarios: {len(chosen)} | "
          f"fuel: {'all Power*' if fuel_filter is None else ','.join(fuel_filter)}")
    print(f"  wrote {os.path.basename(out_path)}")
    print(f"    Shadow Prices (TS)     {n_ts} rows")
    print(f"    Gen-Weighted US & NA   {n_wtd} rows")
    print(f"    Annual Average         {n_ann} rows")
    print("Done.")


if __name__ == "__main__":
    main()
