#!/usr/bin/env python3
"""
export_dispatch_prices.py

Export power prices from the DISPATCH database in EUR/MWh, mirroring
export_shadow_prices.py so the investment-model shadow prices and the dispatch
prices can be compared sheet-for-sheet.

Source:
    dispatch_balance     hourly Price per Region (already in EUR/MWh - the
                         dispatch summary reports AvgPrice_EURMWh, so NO unit
                         conversion is applied here, unlike the M EUR/PJ duals
                         in the investment model).
    dispatch_generation  hourly generation by Region/Technology -> summed to
                         Region/Hour as the generation weights.

Output sheets:
    Prices (hourly)      Scenario | Region | Year | Hour | Price_EUR_MWh
    Gen-Weighted US & NA Scenario | Aggregate | Year | GenWtdPrice_EUR_MWh
                         P = sum(price_h,region * gen_h,region)/sum(gen_h,region)
    Annual Average       Scenario | Region | Year | AvgPrice_EUR_MWh
                         (simple mean over hours - NOT weighted)

Run:
    python export_dispatch_prices.py
    python export_dispatch_prices.py --scenarios all -o dispatch_prices.xlsx
"""

import os
import sys
import argparse
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(SCRIPT_DIR, "..", "Results")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "Output")
DISPATCH_DB = os.path.join(RESULTS_DIR, "genesysmod_dispatch_results.duckdb")

PRICE_TABLE = "dispatch_balance"       # hourly Price per Region (EUR/MWh)
GEN_TABLE = "dispatch_generation"      # hourly generation by Region/Technology

US_REGIONS = ["California", "ERCOT", "MISO", "NewEngland", "NewYork",
              "PJM", "SERC", "SPP", "WECC"]
CANADA = "Canada"
AGGREGATES = {
    "Total US": set(US_REGIONS),
    "Total North America": set(US_REGIONS) | {CANADA},
}
REGION_ORDER = {r: i for i, r in enumerate(US_REGIONS + [CANADA])}

# --------------------------------------------------------------------------- #
# Scenario selection (multi)
# --------------------------------------------------------------------------- #
def distinct_scenarios(con):
    return sorted({(s, p, ps) for s, p, ps in con.execute(
        f"SELECT DISTINCT Scenario, Pathway, PathwayScenario FROM {PRICE_TABLE}"
    ).fetchall()})


def scenario_clause(combo):
    s, p, ps = combo
    esc = lambda x: str(x).replace("'", "''")
    return (f"Scenario='{esc(s)}' AND Pathway='{esc(p)}' "
            f"AND PathwayScenario='{esc(ps)}'")


def _match(scenarios, token):
    token = token.strip()
    if token.isdigit() and 1 <= int(token) <= len(scenarios):
        return scenarios[int(token) - 1]
    for c in scenarios:
        if token == c[0] or token == c[2]:
            return c
    return None


def choose_scenarios(scenarios, preset=None):
    if not scenarios:
        sys.exit("[error] no scenarios in dispatch price table.")
    if preset:
        if preset.strip().lower() == "all":
            return scenarios
        chosen = [_match(scenarios, t) for t in preset.split(",")]
        if any(c is None for c in chosen):
            sys.exit(f"[error] unknown scenario in --scenarios '{preset}'.")
        return chosen
    if len(scenarios) == 1:
        print(f"  single scenario -> {scenarios[0][0]}")
        return scenarios
    print("\nAvailable scenarios:")
    for i, (s, p, ps) in enumerate(scenarios, 1):
        print(f"  [{i}] Scenario={s} | PathwayScenario={ps}")
    while True:
        try:
            raw = input("Select scenarios (comma-separated, or 'all'): ").strip()
        except EOFError:
            sys.exit("[error] no input. Pass --scenarios.")
        if raw.lower() == "all":
            return scenarios
        chosen = [_match(scenarios, t) for t in raw.split(",") if t.strip()]
        if chosen and all(c is not None for c in chosen):
            seen, out = set(), []
            for c in chosen:
                if c not in seen:
                    seen.add(c); out.append(c)
            return out
        print("  invalid selection, try again.")


# --------------------------------------------------------------------------- #
# Data
# --------------------------------------------------------------------------- #
def fetch_prices(con, where):
    """(Scenario, Region, Year, Hour, Price_EUR_MWh) - no unit conversion."""
    rows = con.execute(
        f"SELECT Scenario, Region, Year, Hour, Price FROM {PRICE_TABLE} "
        f"WHERE {where} AND Price IS NOT NULL"
    ).fetchall()
    return [(s, reg, int(y), h, p) for s, reg, y, h, p in rows]


def fetch_generation(con, where):
    """Generation weights keyed (scen, region, year, hour)."""
    rows = con.execute(
        f"SELECT Scenario, Region, Year, Hour, SUM(Value) FROM {GEN_TABLE} "
        f"WHERE {where} GROUP BY Scenario, Region, Year, Hour"
    ).fetchall()
    gen = {}
    for s, reg, y, h, v in rows:
        if v is not None:
            gen[(s, reg, int(y), h)] = v
    return gen


# --------------------------------------------------------------------------- #
# Workbook
# --------------------------------------------------------------------------- #
HDR_FILL = PatternFill("solid", fgColor="7F3F1F")   # brown, to visually differ
HDR_FONT = Font(bold=True, color="FFFFFF")


def _sheet(ws, headers, rows, num_cols):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(r)
    for c in num_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    for c in range(1, len(headers) + 1):
        w = max(len(str(headers[c - 1])) + 2,
                max((len(str(ws.cell(r, c).value))
                     for r in range(2, min(ws.max_row, 200) + 1)), default=8) + 2)
        ws.column_dimensions[get_column_letter(c)].width = min(w, 40)


def build_workbook(prices, gen, out_path):
    wb = Workbook()
    wb.remove(wb.active)

    hp = sorted(prices, key=lambda r: (r[0], REGION_ORDER.get(r[1], 99), r[2], r[3]))
    hp_out = [(s, reg, y, h, round(p, 4)) for (s, reg, y, h, p) in hp]

    # per-region annual mean (over hours)
    acc = {}
    for s, reg, y, h, p in prices:
        a = acc.setdefault((s, reg, y), [0.0, 0]); a[0] += p; a[1] += 1
    ann = [(s, reg, y, round(v[0] / v[1], 4)) for (s, reg, y), v in acc.items()]
    ann.sort(key=lambda r: (r[0], REGION_ORDER.get(r[1], 99), r[2]))

    # generation-weighted Total US / Total NA
    wacc = {}
    for s, reg, y, h, p in prices:
        g = gen.get((s, reg, y, h))
        if g is None:
            continue
        for aggname, members in AGGREGATES.items():
            if reg in members:
                a = wacc.setdefault((s, aggname, y), [0.0, 0.0])
                a[0] += p * g; a[1] += g
    agg_order = {"Total US": 0, "Total North America": 1}
    wtd = [(s, agg, y, round(w[0] / w[1], 4))
           for (s, agg, y), w in wacc.items() if w[1] > 0]
    wtd.sort(key=lambda r: (r[0], agg_order.get(r[1], 9), r[2]))

    _sheet(wb.create_sheet("Prices (hourly)"),
           ["Scenario", "Region", "Year", "Hour", "Price_EUR_MWh"],
           hp_out, num_cols=[5])
    _sheet(wb.create_sheet("Gen-Weighted US & NA"),
           ["Scenario", "Aggregate", "Year", "GenWtdPrice_EUR_MWh"],
           wtd, num_cols=[4])
    _sheet(wb.create_sheet("Annual Average"),
           ["Scenario", "Region", "Year", "AvgPrice_EUR_MWh"],
           ann, num_cols=[4])

    wsw = wb["Gen-Weighted US & NA"]
    n1 = wsw.cell(wsw.max_row + 2, 1)
    n1.value = ("Generation-weighted over hours & member regions, weights = "
                "dispatch_generation. Dispatch prices are already EUR/MWh "
                "(no unit conversion). Compare with the investment-model file "
                "'Shadow_Prices_Power.xlsx'.")
    n1.font = Font(italic=True, size=9, color="808080")

    ws = wb["Annual Average"]
    note = ws.cell(ws.max_row + 2, 1)
    note.value = "Note: simple mean over hours (NOT weighted)."
    note.font = Font(italic=True, size=9, color="808080")

    wb.save(out_path)
    return len(hp_out), len(wtd), len(ann)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(
        description="Export dispatch power prices (EUR/MWh) + gen-weighted US/NA.")
    ap.add_argument("--scenarios", default=os.environ.get("GMOD_SCENARIOS"),
                    help="'all' or comma-separated Scenario/PathwayScenario names.")
    ap.add_argument("-o", "--output", default=None)
    args = ap.parse_args()

    if not os.path.exists(DISPATCH_DB):
        sys.exit(f"Missing dispatch database: {DISPATCH_DB}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    con = duckdb.connect(DISPATCH_DB, read_only=True)
    chosen = choose_scenarios(distinct_scenarios(con), preset=args.scenarios)
    where = "(" + " OR ".join(f"({scenario_clause(c)})" for c in chosen) + ")"

    prices = fetch_prices(con, where)
    gen = fetch_generation(con, where)
    con.close()
    if not prices:
        sys.exit("[error] no price rows for the chosen scenarios.")

    out_path = args.output or os.path.join(OUTPUT_DIR, "Dispatch_Prices_Power.xlsx")
    n_hp, n_wtd, n_ann = build_workbook(prices, gen, out_path)

    print(f"\n  scenarios: {len(chosen)}")
    print(f"  wrote {os.path.basename(out_path)}")
    print(f"    Prices (hourly)        {n_hp} rows")
    print(f"    Gen-Weighted US & NA   {n_wtd} rows")
    print(f"    Annual Average         {n_ann} rows")
    print("Done.")


if __name__ == "__main__":
    main()
