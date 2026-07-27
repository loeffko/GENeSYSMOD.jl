#!/usr/bin/env python3
"""
dispatch_gen_diff.py

Difference in annual generation (sensitivity - base) by technology, for one
target year, taken from the DISPATCH database.  Conventional generation comes
from dispatch_gen_annual (Generation_GWh); storage (BESS/LDES/PHES) comes from
dispatch_storage as GROSS DISCHARGE only (Discharge summed over hours/regions,
charge ignored).  All values in GWh.

Interactively (or via CLI) select:
    a) the base case               (one scenario)
    b) one or more sensitivities   (multi-select)
    c) a target year

Output: an Excel file with technologies down the rows and one column per
sensitivity, values = (sensitivity - base) generation in GWh.  Cells that come
out to zero are left blank.  Columns carry a two-row header (group / label).

Run:
    python dispatch_gen_diff.py
    python dispatch_gen_diff.py --base 73_ramping_invlimit \
        --sensitivities "BESS_high,BESS_low,Demand_high" --year 2040
    python dispatch_gen_diff.py --base BASE --sensitivities all --year 2040 \
        --labels "BESS_high=BESS|BESS+;BESS_low=BESS|BESS -"
"""

import os
import sys
import argparse
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(SCRIPT_DIR, "..", "Results")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "Output")
DISPATCH_DB = os.path.join(RESULTS_DIR, "genesysmod_dispatch_results.duckdb")

GEN_TABLE = "dispatch_gen_annual"          # Year, Technology, Generation_GWh
STORAGE_TABLE = "dispatch_storage"         # hourly Charge/Discharge/SoC per S_*
VALUE_UNIT = "GWh"                          # output in GWh (native table unit)

# Storage discharge -> ThinkCell storage categories
STORAGE_MAP = {
    "S_Battery_Li-Ion": "BESS",
    "S_Battery_Redox": "LDES",
    "S_CAES": "LDES",
    "S_PHS": "PHES",
}
STORAGE_CATS = {"BESS", "LDES", "PHES"}

# --------------------------------------------------------------------------- #
# Technology mapping (same P_/D_ codes as the results DB; ThinkCell 13-cat)
# --------------------------------------------------------------------------- #
THINKCELL_MAP = {
    "P_Coal_Hardcoal": "Coal", "P_Coal_Hardcoal_CCS": "Coal",
    "P_Coal_Lignite": "Coal", "P_Coal_Lignite_CCS": "Coal",
    "CHP_Coal_Hardcoal": "Coal", "CHP_Coal_Hardcoal_CCS": "Coal",
    "CHP_Coal_Lignite": "Coal", "CHP_Coal_Lignite_CCS": "Coal",
    "P_Gas_CCGT": "Gas", "P_Gas_OCGT": "Gas", "P_Gas_Engines": "Gas",
    "P_Gas_Steam": "Gas", "P_Gas_CCGT_Residual": "Gas", "P_SOFC": "Gas",
    "CHP_Gas_CCGT_Natural": "Gas", "CHP_Gas_CCGT_Biogas": "Gas",
    "P_Gas_CCS": "Gas CCS", "CHP_Gas_CCGT_Natural_CCS": "Gas CCS",
    "CHP_Gas_CCGT_Biogas_CCS": "Gas CCS",
    "P_Oil": "Other", "CHP_Oil": "Other",
    "P_Geothermal": "Other", "P_EGS_R1": "Other", "P_EGS_R2": "Other",
    "P_EGS_R3": "Other", "P_EGS_R4": "Other",
    "P_Ocean": "Other", "CHP_WasteToEnergy": "Other",
    "P_H2_OCGT": "Other", "CHP_Hydrogen_FuelCell": "Other",
    "P_Nuclear": "Nuclear", "P_Nuclear_SMR": "Nuclear",
    "P_Hydro_Reservoir": "Hydro", "P_Hydro_RoR": "Hydro",
    "P_Biomass": "Biomass", "P_Biomass_CCS": "Biomass",
    "CHP_Biomass_Solid": "Biomass", "CHP_Biomass_Solid_CCS": "Biomass",
    "P_PV_Rooftop_Commercial": "Solar", "P_PV_Rooftop_Residential": "Solar",
    "P_PV_Utility_Avg": "Solar", "P_PV_Utility_Inf": "Solar",
    "P_PV_Utility_Opt": "Solar", "P_PV_Utility_Tracking": "Solar",
    "P_CSP": "Solar",
    "P_Wind_Onshore_Avg": "Wind Onshore", "P_Wind_Onshore_Inf": "Wind Onshore",
    "P_Wind_Onshore_Opt": "Wind Onshore",
    "P_Wind_Offshore_Shallow": "Wind Offshore",
    "P_Wind_Offshore_Transitional": "Wind Offshore",
    "P_Wind_Offshore_Deep": "Wind Offshore",
    "D_Battery_Li-Ion": "BESS",
    "D_Battery_Redox": "LDES", "D_CAES": "LDES",
    "D_PHS": "PHES",
}
TECH_ORDER = ["Coal", "Gas", "Gas CCS", "Other", "Nuclear", "Hydro",
              "Biomass", "Solar", "Wind Onshore", "Wind Offshore",
              "BESS", "LDES", "PHES"]

# --------------------------------------------------------------------------- #
# Scenario / year selection
# --------------------------------------------------------------------------- #
def distinct_scenarios(con):
    return sorted({(s, p, ps) for s, p, ps in con.execute(
        f"SELECT DISTINCT Scenario, Pathway, PathwayScenario FROM {GEN_TABLE}"
    ).fetchall()})


def scenario_clause(combo):
    s, p, ps = combo
    esc = lambda x: str(x).replace("'", "''")
    return (f"Scenario='{esc(s)}' AND Pathway='{esc(p)}' "
            f"AND PathwayScenario='{esc(ps)}'")


def _match(scenarios, token):
    token = token.strip()
    if token.isdigit() and 1 <= int(token) <= len(scenarios):
        return scenarios[int(token) - 1]
    for c in scenarios:
        if token == c[0] or token == c[2]:
            return c
    return None


def _print_list(scenarios, header):
    print(f"\n{header}")
    for i, (s, p, ps) in enumerate(scenarios, 1):
        print(f"  [{i}] Scenario={s} | PathwayScenario={ps}")


def choose_base(scenarios, preset=None):
    if preset:
        c = _match(scenarios, preset)
        if not c:
            sys.exit(f"[error] base '{preset}' not found.")
        return c
    if len(scenarios) == 1:
        print(f"  base case (only scenario) -> {scenarios[0][0]}")
        return scenarios[0]
    _print_list(scenarios, "Available scenarios:")
    while True:
        try:
            raw = input("Select BASE case (number or name): ").strip()
        except EOFError:
            sys.exit("[error] no input for base. Pass --base.")
        c = _match(scenarios, raw)
        if c:
            return c
        print("  invalid, try again.")


def choose_sensitivities(scenarios, base, preset=None):
    pool = [c for c in scenarios if c != base]
    if not pool:
        sys.exit("[error] no scenarios left for sensitivities (only base).")
    if preset:
        if preset.strip().lower() == "all":
            return pool
        chosen = [_match(pool, t) for t in preset.split(",")]
        if any(c is None for c in chosen):
            sys.exit(f"[error] unknown sensitivity in '{preset}'.")
        return chosen
    _print_list(pool, "Sensitivities (base excluded):")
    while True:
        try:
            raw = input("Select SENSITIVITIES (comma-separated, or 'all'): "
                        ).strip()
        except EOFError:
            sys.exit("[error] no input for sensitivities. Pass --sensitivities.")
        if raw.lower() == "all":
            return pool
        chosen = [_match(pool, t) for t in raw.split(",") if t.strip()]
        if chosen and all(c is not None for c in chosen):
            seen, out = set(), []
            for c in chosen:
                if c not in seen:
                    seen.add(c); out.append(c)
            return out
        print("  invalid selection, try again.")


def choose_year(con, preset=None):
    years = sorted({r[0] for r in
                    con.execute(f"SELECT DISTINCT Year FROM {GEN_TABLE}").fetchall()})
    if preset is not None:
        y = int(preset)
        if y not in years:
            sys.exit(f"[error] year {y} not in dispatch data {years}.")
        return y
    if len(years) == 1:
        print(f"  target year (only year) -> {years[0]}")
        return years[0]
    print(f"\nAvailable years: {years}")
    while True:
        try:
            raw = input("Select TARGET year: ").strip()
        except EOFError:
            sys.exit("[error] no input for year. Pass --year.")
        if raw.isdigit() and int(raw) in years:
            return int(raw)
        print("  invalid year, try again.")


# --------------------------------------------------------------------------- #
# Generation / storage by mapped category (GWh) for one scenario+year
# --------------------------------------------------------------------------- #
def gen_by_cat(con, where, year):
    """Non-storage generation from the gen table (GWh). Storage categories
    (BESS/LDES/PHES) are excluded here and taken from the storage table."""
    rows = con.execute(
        f"SELECT Technology, SUM(Generation_GWh) FROM {GEN_TABLE} "
        f"WHERE {where} AND Year=? GROUP BY Technology", [year]
    ).fetchall()
    out = {}
    for tech, val in rows:
        cat = THINKCELL_MAP.get(tech)
        if cat is None or val is None or cat in STORAGE_CATS:
            continue
        out[cat] = out.get(cat, 0.0) + val
    return out


def storage_by_cat(con, where, year):
    """Gross storage discharge (no net-charge) from the storage table (GWh),
    summed over hours & regions, mapped S_* -> BESS/LDES/PHES."""
    out = {}
    try:
        rows = con.execute(
            f"SELECT Storage, SUM(Discharge) FROM {STORAGE_TABLE} "
            f"WHERE {where} AND Year=? GROUP BY Storage", [year]
        ).fetchall()
    except Exception as e:
        print(f"  [warn] storage table '{STORAGE_TABLE}' unavailable: {e}")
        return out
    for storage, val in rows:
        cat = STORAGE_MAP.get(storage)
        if cat is None or val is None:
            continue
        out[cat] = out.get(cat, 0.0) + val
    return out


def totals_by_cat(con, where, year):
    """Combined generation (gen table) + gross storage discharge (sto table)."""
    d = gen_by_cat(con, where, year)
    for cat, v in storage_by_cat(con, where, year).items():
        d[cat] = d.get(cat, 0.0) + v
    return d


# --------------------------------------------------------------------------- #
# Labels for the two-row column header
# --------------------------------------------------------------------------- #
def resolve_labels(sensitivities, labels_arg):
    """Return list of (group, sublabel) per sensitivity."""
    preset = {}
    if labels_arg:
        for part in labels_arg.split(";"):
            if "=" not in part:
                continue
            scen, lab = part.split("=", 1)
            grp, sub = (lab.split("|", 1) + [""])[:2] if "|" in lab else ("", lab)
            preset[scen.strip()] = (grp.strip(), sub.strip())
    out = []
    for combo in sensitivities:
        scen = combo[0]
        if scen in preset:
            out.append(preset[scen])
        elif combo[2] in preset:
            out.append(preset[combo[2]])
        else:
            out.append(("", scen))          # default: group blank, sub=Scenario
    return out


# --------------------------------------------------------------------------- #
# Workbook
# --------------------------------------------------------------------------- #
def write_workbook(out_path, base, sensitivities, labels, year,
                   base_gen, sens_gens):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Gen Diff {year}"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(bottom=thin, right=thin)
    grp_fill = PatternFill("solid", fgColor="D9E1F2")
    sub_fill = PatternFill("solid", fgColor="F2F2F2")

    # corner label
    ws.cell(2, 1).value = f"\u0394 Generation {VALUE_UNIT} ({year}) vs {base[0]}"
    ws.cell(2, 1).font = Font(bold=True, size=9, italic=True)

    # headers + values
    for j, combo in enumerate(sensitivities):
        col = 2 + j
        grp, sub = labels[j]
        gcell = ws.cell(1, col); gcell.value = grp
        gcell.font = Font(bold=True); gcell.fill = grp_fill
        gcell.alignment = Alignment(horizontal="center"); gcell.border = border
        scell = ws.cell(2, col); scell.value = sub
        scell.font = Font(bold=True); scell.fill = sub_fill
        scell.alignment = Alignment(horizontal="center"); scell.border = border

    for i, cat in enumerate(TECH_ORDER):
        row = 3 + i
        c = ws.cell(row, 1); c.value = cat; c.font = Font(bold=True)
        for j, combo in enumerate(sensitivities):
            diff = (sens_gens[j].get(cat, 0.0) - base_gen.get(cat, 0.0))
            v = round(diff)
            cell = ws.cell(row, 2 + j)
            cell.value = v if v != 0 else None      # blank when zero
            cell.number_format = "#,##0"

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 15
    for j in range(len(sensitivities)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 12

    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(
        description="Generation difference (sensitivity - base) from dispatch DB.")
    ap.add_argument("--base", default=None, help="Base-case Scenario name.")
    ap.add_argument("--sensitivities", default=None,
                    help="'all' or comma-separated Scenario names.")
    ap.add_argument("--year", default=None, help="Target year.")
    ap.add_argument("--labels", default=None,
                    help="Optional 'scen=Group|Sub;...' two-row header labels.")
    ap.add_argument("-o", "--output", default=None)
    args = ap.parse_args()

    if not os.path.exists(DISPATCH_DB):
        sys.exit(f"Missing dispatch database: {DISPATCH_DB}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    con = duckdb.connect(DISPATCH_DB, read_only=True)
    scenarios = distinct_scenarios(con)

    base = choose_base(scenarios, args.base)
    sensitivities = choose_sensitivities(scenarios, base, args.sensitivities)
    year = choose_year(con, args.year)

    # optional interactive custom labels
    labels_arg = args.labels
    if labels_arg is None and sys.stdin.isatty():
        ans = input("\nEnter custom two-row column labels? [y/N]: ").strip().lower()
        if ans == "y":
            parts = []
            for combo in sensitivities:
                grp = input(f"  group label for '{combo[0]}': ").strip()
                sub = input(f"  sub-label  for '{combo[0]}': ").strip() or combo[0]
                parts.append(f"{combo[0]}={grp}|{sub}")
            labels_arg = ";".join(parts)
    labels = resolve_labels(sensitivities, labels_arg)

    base_gen = totals_by_cat(con, scenario_clause(base), year)
    sens_gens = [totals_by_cat(con, scenario_clause(c), year) for c in sensitivities]
    con.close()

    out_path = args.output or os.path.join(
        OUTPUT_DIR, f"Dispatch_GenDiff_{year}.xlsx")
    write_workbook(out_path, base, sensitivities, labels, year,
                   base_gen, sens_gens)

    print(f"\n  base: {base[0]}")
    print(f"  sensitivities ({len(sensitivities)}): "
          f"{', '.join(c[0] for c in sensitivities)}")
    print(f"  year: {year}")
    print(f"  wrote {os.path.basename(out_path)}")
    print("Done.")


if __name__ == "__main__":
    main()
