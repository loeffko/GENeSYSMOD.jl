#!/usr/bin/env python3
"""
kpi_summary.py

Combine selected GENeSYS-MOD KPIs into ONE Excel workbook, aggregated to
country level (US = 9 US regions, CA = Canada), long format with a Scenario
column so multiple scenarios stack in the same sheets.

Sheets (one KPI each):
    Installed Capacity      Scenario | Country | Technology | Year | Capacity_GW
    Power Generation        Scenario | Country | Technology | Year | Generation_TWh
    BESS Capacity           Scenario | Country | Year | Power_GW | Energy_GWh
    Interconnector Capacity Scenario | From | To | Year | Capacity_GW   (both directions, US<->CA)

Sources (results DB):
    output_capacity                 Type='TotalCapacity'      -> capacity / BESS power (GW)
    output_annual_production        Type='Production'         -> generation (PJ -> TWh)
    raw_TotalStorageCapacityAnnual  Storage='S_...'           -> BESS energy (GWh)
    output_trade                    Type='Transmissions Capacity' -> interconnectors (GW)

Run:  python kpi_summary.py                 (interactive multi-select)
      python kpi_summary.py --scenarios all
      python kpi_summary.py --scenarios "244_invLimit_ramping_v01,ALT_run_v02"
"""

import os
import sys
import argparse
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(SCRIPT_DIR, "..", "Results")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "Output")
RESULTS_DB = os.path.join(RESULTS_DIR, "genesysmod_db.duckdb")

# --------------------------------------------------------------------------- #
# Regions -> country
# --------------------------------------------------------------------------- #
US_REGIONS = ["California", "ERCOT", "MISO", "NewEngland", "NewYork",
              "PJM", "SERC", "SPP", "WECC"]
CANADA = "Canada"


def country_of(region):
    if region == CANADA:
        return "CA"
    if region in US_REGIONS:
        return "US"
    return None


# --------------------------------------------------------------------------- #
# Technology mapping (ThinkCell 13-cat scheme; storage kept)
# --------------------------------------------------------------------------- #
THINKCELL_MAP = {
    "P_Coal_Hardcoal": "Coal", "P_Coal_Hardcoal_CCS": "Coal",
    "P_Coal_Lignite": "Coal", "P_Coal_Lignite_CCS": "Coal",
    "CHP_Coal_Hardcoal": "Coal", "CHP_Coal_Hardcoal_CCS": "Coal",
    "CHP_Coal_Lignite": "Coal", "CHP_Coal_Lignite_CCS": "Coal",
    "P_Gas_CCGT": "Gas", "P_Gas_OCGT": "Gas", "P_Gas_Engines": "Gas",
    "P_Gas_Steam": "Gas", "P_Gas_CCGT_Residual": "Gas", "P_SOFC": "Gas",
    "CHP_Gas_CCGT_Natural": "Gas", "CHP_Gas_CCGT_Biogas": "Gas",
    "P_Gas_CCS": "Gas CCS", "CHP_Gas_CCGT_Natural_CCS": "Gas CCS",
    "CHP_Gas_CCGT_Biogas_CCS": "Gas CCS",
    "P_Oil": "Other", "CHP_Oil": "Other",
    "P_Geothermal": "Other", "P_EGS_R1": "Other", "P_EGS_R2": "Other",
    "P_EGS_R3": "Other", "P_EGS_R4": "Other",
    "P_Ocean": "Other", "CHP_WasteToEnergy": "Other",
    "P_H2_OCGT": "Other", "CHP_Hydrogen_FuelCell": "Other",
    "P_Nuclear": "Nuclear", "P_Nuclear_SMR": "Nuclear",
    "P_Hydro_Reservoir": "Hydro", "P_Hydro_RoR": "Hydro",
    "P_Biomass": "Biomass", "P_Biomass_CCS": "Biomass",
    "CHP_Biomass_Solid": "Biomass", "CHP_Biomass_Solid_CCS": "Biomass",
    "P_PV_Rooftop_Commercial": "Solar", "P_PV_Rooftop_Residential": "Solar",
    "P_PV_Utility_Avg": "Solar", "P_PV_Utility_Inf": "Solar",
    "P_PV_Utility_Opt": "Solar", "P_PV_Utility_Tracking": "Solar",
    "P_CSP": "Solar",
    "P_Wind_Onshore_Avg": "Wind Onshore", "P_Wind_Onshore_Inf": "Wind Onshore",
    "P_Wind_Onshore_Opt": "Wind Onshore",
    "P_Wind_Offshore_Shallow": "Wind Offshore",
    "P_Wind_Offshore_Transitional": "Wind Offshore",
    "P_Wind_Offshore_Deep": "Wind Offshore",
    "D_Battery_Li-Ion": "BESS",
    "D_Battery_Redox": "LDES", "D_CAES": "LDES",
    "D_PHS": "PHES",
}
TECH_ORDER = ["Coal", "Gas", "Gas CCS", "Other", "Nuclear", "Hydro",
              "Biomass", "Solar", "Wind Onshore", "Wind Offshore",
              "BESS", "LDES", "PHES"]

BESS_POWER_TECH = "D_Battery_Li-Ion"       # GW  (output_capacity TotalCapacity)
BESS_ENERGY_STORAGE = "S_Battery_Li-Ion"   # GWh (raw_TotalStorageCapacityAnnual)

PJ_TO_TWH = 1.0 / 3.6
PJ_TO_GWH = 1000.0 / 3.6      # storage energy (raw_TotalStorageCapacityAnnual) is in PJ

# --------------------------------------------------------------------------- #
# Scenario selection (multi-select)
# --------------------------------------------------------------------------- #
def distinct_scenarios(con, tables):
    combos = set()
    for t in tables:
        try:
            for s, p, ps in con.execute(
                f"SELECT DISTINCT Scenario, Pathway, PathwayScenario FROM {t}"
            ).fetchall():
                combos.add((s, p, ps))
        except Exception:
            pass
    return sorted(combos)


def scenario_clause(combo):
    s, p, ps = combo
    esc = lambda x: str(x).replace("'", "''")
    return (f"Scenario='{esc(s)}' AND Pathway='{esc(p)}' "
            f"AND PathwayScenario='{esc(ps)}'")


def choose_scenarios(scenarios, preset=None):
    """Return a list of chosen combos (>=1). Multi-select."""
    if not scenarios:
        sys.exit("[error] no scenarios found in results DB.")

    def match(token):
        token = token.strip()
        if token.isdigit() and 1 <= int(token) <= len(scenarios):
            return scenarios[int(token) - 1]
        for c in scenarios:
            if token == c[0] or token == c[2]:
                return c
        return None

    if preset:
        if preset.strip().lower() == "all":
            return scenarios
        chosen = [match(tok) for tok in preset.split(",")]
        if any(c is None for c in chosen):
            sys.exit(f"[error] unknown scenario in --scenarios '{preset}'.")
        return chosen

    if len(scenarios) == 1:
        s, _, ps = scenarios[0]
        print(f"  single scenario -> {s} | {ps}")
        return scenarios

    print("\nAvailable scenarios:")
    for i, (s, p, ps) in enumerate(scenarios, 1):
        print(f"  [{i}] Scenario={s} | Pathway={p} | PathwayScenario={ps}")
    while True:
        try:
            raw = input("Select scenarios (comma-separated numbers/names, "
                        "or 'all'): ").strip()
        except EOFError:
            sys.exit("[error] no input. Re-run interactively or pass "
                     "--scenarios.")
        if raw.lower() == "all":
            return scenarios
        chosen = [match(tok) for tok in raw.split(",") if tok.strip()]
        if chosen and all(c is not None for c in chosen):
            # de-dup, preserve order
            seen, out = set(), []
            for c in chosen:
                if c not in seen:
                    seen.add(c)
                    out.append(c)
            return out
        print("  invalid selection, try again.")


# --------------------------------------------------------------------------- #
# KPI computation for one scenario
# --------------------------------------------------------------------------- #
def kpi_capacity(con, where):
    rows = con.execute(
        f"SELECT Region, Technology, Year, Value FROM output_capacity "
        f"WHERE {where} AND Type='TotalCapacity'"
    ).fetchall()
    agg = {}  # (country, cat, year) -> GW
    for region, tech, year, val in rows:
        c = country_of(region)
        cat = THINKCELL_MAP.get(tech)
        if c is None or cat is None:
            continue
        agg[(c, cat, year)] = agg.get((c, cat, year), 0.0) + val
    return agg


def kpi_generation(con, where):
    rows = con.execute(
        f"SELECT Region, Technology, Year, Value FROM output_annual_production "
        f"WHERE {where} AND Type='Production'"
    ).fetchall()
    agg = {}  # (country, cat, year) -> TWh
    for region, tech, year, val in rows:
        c = country_of(region)
        cat = THINKCELL_MAP.get(tech)
        if c is None or cat is None:
            continue
        agg[(c, cat, year)] = agg.get((c, cat, year), 0.0) + val * PJ_TO_TWH
    return agg


def kpi_bess(con, where):
    power = con.execute(
        f"SELECT Region, Year, SUM(Value) FROM output_capacity "
        f"WHERE {where} AND Type='TotalCapacity' AND Technology=? "
        f"GROUP BY Region, Year", [BESS_POWER_TECH]
    ).fetchall()
    pw = {}  # (country, year) -> GW
    for region, year, val in power:
        c = country_of(region)
        if c:
            pw[(c, year)] = pw.get((c, year), 0.0) + (val or 0.0)

    en = {}  # (country, year) -> GWh
    try:
        energy = con.execute(
            f"SELECT Region, Year, SUM(Value) FROM raw_TotalStorageCapacityAnnual "
            f"WHERE {where} AND Storage=? GROUP BY Region, Year",
            [BESS_ENERGY_STORAGE]
        ).fetchall()
        for region, year, val in energy:
            c = country_of(region)
            if c:
                # raw_TotalStorageCapacityAnnual is in PJ -> GWh
                en[(c, year)] = en.get((c, year), 0.0) + (val or 0.0) * PJ_TO_GWH
    except Exception as e:
        print(f"  [warn] BESS energy unavailable "
              f"(raw_TotalStorageCapacityAnnual): {e}")
    return pw, en


def kpi_interconnector(con, where):
    """Sum Transmissions Capacity by ORIGIN country: every interconnector
    starting in a US region -> US; every one starting in Canada -> CA.
    Includes inner-US (US->US) links, not just cross-border."""
    rows = con.execute(
        f"SELECT Region, Region2, Year, Value FROM output_trade "
        f"WHERE {where} AND Type='Transmissions Capacity'"
    ).fetchall()
    agg = {}  # (origin_country, year) -> GW
    for a, b, year, val in rows:
        c = country_of(a)          # origin country
        if c is None:
            continue
        agg[(c, year)] = agg.get((c, year), 0.0) + val
    return agg


# --------------------------------------------------------------------------- #
# Workbook writing
# --------------------------------------------------------------------------- #
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(ws, headers, rows, num_cols):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(r)
    # number format
    for c in num_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).number_format = "#,##0.000"
    ws.freeze_panes = "A2"
    # column widths
    for c in range(1, len(headers) + 1):
        width = max(len(str(headers[c - 1])) + 2,
                    max((len(str(ws.cell(r, c).value))
                         for r in range(2, min(ws.max_row, 200) + 1)),
                        default=8) + 2)
        ws.column_dimensions[get_column_letter(c)].width = min(width, 40)


def build_workbook(con, chosen, out_path):
    wb = Workbook()
    wb.remove(wb.active)

    cap_rows, gen_rows, bess_rows, ic_rows = [], [], [], []
    country_order = {"US": 0, "CA": 1}

    for combo in chosen:
        scen = combo[0]                       # Scenario label
        where = scenario_clause(combo)

        cap = kpi_capacity(con, where)
        for (c, cat, y), v in cap.items():
            cap_rows.append((scen, c, cat, y, round(v, 4)))
        gen = kpi_generation(con, where)
        for (c, cat, y), v in gen.items():
            gen_rows.append((scen, c, cat, y, round(v, 4)))
        pw, en = kpi_bess(con, where)
        for (c, y) in sorted(set(pw) | set(en)):
            bess_rows.append((scen, c, y,
                              round(pw.get((c, y), 0.0), 4),
                              round(en.get((c, y), 0.0), 4)))
        ic = kpi_interconnector(con, where)
        for (c, y), v in ic.items():
            ic_rows.append((scen, c, y, round(v, 4)))

    # sort each block: Scenario, Country/From, Technology order, Year
    tord = {t: i for i, t in enumerate(TECH_ORDER)}
    cap_rows.sort(key=lambda r: (r[0], country_order.get(r[1], 9),
                                 tord.get(r[2], 99), r[3]))
    gen_rows.sort(key=lambda r: (r[0], country_order.get(r[1], 9),
                                 tord.get(r[2], 99), r[3]))
    bess_rows.sort(key=lambda r: (r[0], country_order.get(r[1], 9), r[2]))
    ic_rows.sort(key=lambda r: (r[0], country_order.get(r[1], 9), r[2]))

    _write_sheet(wb.create_sheet("Installed Capacity"),
                 ["Scenario", "Country", "Technology", "Year", "Capacity_GW"],
                 cap_rows, num_cols=[5])
    _write_sheet(wb.create_sheet("Power Generation"),
                 ["Scenario", "Country", "Technology", "Year", "Generation_TWh"],
                 gen_rows, num_cols=[5])
    _write_sheet(wb.create_sheet("BESS Capacity"),
                 ["Scenario", "Country", "Year", "Power_GW", "Energy_GWh"],
                 bess_rows, num_cols=[4, 5])
    _write_sheet(wb.create_sheet("Interconnector Capacity"),
                 ["Scenario", "Country", "Year", "Capacity_GW"],
                 ic_rows, num_cols=[4])

    wb.save(out_path)
    return len(cap_rows), len(gen_rows), len(bess_rows), len(ic_rows)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(
        description="Combine GENeSYS-MOD KPIs (US/CA) into one workbook.")
    ap.add_argument("--scenarios", default=os.environ.get("GMOD_SCENARIOS"),
                    help="'all' or comma-separated Scenario/PathwayScenario "
                         "names to skip the prompt.")
    ap.add_argument("-o", "--output", default=None,
                    help="Output xlsx path (default Output/KPI_Summary_NA.xlsx)")
    args = ap.parse_args()

    if not os.path.exists(RESULTS_DB):
        sys.exit(f"Missing database: {RESULTS_DB}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = args.output or os.path.join(OUTPUT_DIR, "KPI_Summary_NA.xlsx")

    con = duckdb.connect(RESULTS_DB, read_only=True)
    scenarios = distinct_scenarios(
        con, ["output_capacity", "output_annual_production", "output_trade"])
    chosen = choose_scenarios(scenarios, preset=args.scenarios)

    print(f"\n  scenarios selected: {len(chosen)}")
    for c in chosen:
        print(f"    - {c[0]} | {c[2]}")

    nc, ng, nb, ni = build_workbook(con, chosen, out_path)
    con.close()
    print(f"\n  wrote {os.path.basename(out_path)}")
    print(f"    Installed Capacity      {nc} rows")
    print(f"    Power Generation        {ng} rows")
    print(f"    BESS Capacity           {nb} rows")
    print(f"    Interconnector Capacity {ni} rows")
    print("Done.")


if __name__ == "__main__":
    main()
