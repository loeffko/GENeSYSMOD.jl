#!/usr/bin/env python3
"""
convert_genesysmod_na.py

Convert GENeSYS-MOD North-America results into the FEL output templates.

Folder layout (relative to this script):
    Scripts/convert_genesysmod_na.py        <- this file
    Scripts/templates/                      <- input templates
    Results/genesysmod_results_db.duckdb    <- capacity / production / trade
    Results/genesysmod_dispatch_results.duckdb  <- dispatch summary (prices)
    Output/                                 <- written here (created if missing)

Outputs:
    FEL_BaseCase_NA_Generation_ThinkCell.xlsx
    FEL_BaseCase_NA_Capacities_ThinkCell.xlsx
    260618_Base_case_results_beyond_Capacity_and_generation_v3.xlsx  (NAM tabs filled)
    2026_FEL_ElectricityGenerationByTechnology_Input_v03-<yymmdd>.xlsx (region tabs appended)

Run:  python convert_genesysmod_na.py
"""

import os
import sys
import datetime
import duckdb
from openpyxl import load_workbook

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(SCRIPT_DIR, "templates")
RESULTS_DIR = os.path.join(SCRIPT_DIR, "..", "Results")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "Output")

RESULTS_DB = os.path.join(RESULTS_DIR, "genesysmod_db.duckdb")
DISPATCH_DB = os.path.join(RESULTS_DIR, "genesysmod_dispatch_results.duckdb")

TPL_GEN = os.path.join(TEMPLATE_DIR, "FEL_BaseCase_EU_Generation_ThinkCell.xlsx")
TPL_CAP = os.path.join(TEMPLATE_DIR, "FEL_BaseCase_EU_Capacities_ThinkCell.xlsx")
TPL_BEYOND = os.path.join(TEMPLATE_DIR,
                          "260618_Base_case_results_beyond_Capacity_and_generation_v3.xlsx")
TPL_FEL_INPUT = os.path.join(TEMPLATE_DIR,
                             "2026_FEL_ElectricityGenerationByTechnology_Input_v02-260619.xlsx")

# --------------------------------------------------------------------------- #
# Regions
# --------------------------------------------------------------------------- #
US_REGIONS = ["California", "ERCOT", "MISO", "NewEngland", "NewYork",
              "PJM", "SERC", "SPP", "WECC"]          # the 9 US regions
CANADA = "Canada"
TOTAL_US = "Total US"
TOTAL_NA = "Total NA"          # US (9 regions) + Canada

DISPLAY = {
    "California": "California", "ERCOT": "ERCOT", "MISO": "MISO",
    "NewEngland": "New England", "NewYork": "New York", "PJM": "PJM",
    "SERC": "SERC", "SPP": "SPP", "WECC": "WECC", "Canada": "Canada",
}


def disp_name(key):
    """Sheet/label name for a region key (totals keep their own name)."""
    return DISPLAY.get(key, key)


# Tab order for files with totals (gen/cap ThinkCell, FEL input)
ORDER_WITH_TOTAL = [TOTAL_NA, TOTAL_US] + US_REGIONS + [CANADA]
# Region order for file3 (no totals)
ORDER_NO_TOTAL = US_REGIONS + [CANADA]

# --------------------------------------------------------------------------- #
# Unit conversions
# --------------------------------------------------------------------------- #
PJ_TO_TWH = 1.0 / 3.6          # 1 PJ = 0.277778 TWh
PJ_TO_GWH = 1000.0 / 3.6       # 1 PJ = 277.778 GWh
PJ_TO_MWH = 1_000_000.0 / 3.6  # 1 PJ = 277 777.8 MWh
GW_TO_MW = 1000.0

# --------------------------------------------------------------------------- #
# Technology mapping
# --------------------------------------------------------------------------- #
# ThinkCell scheme (13 categories) -- generation & capacity
THINKCELL_MAP = {
    # Coal
    "P_Coal_Hardcoal": "Coal", "P_Coal_Hardcoal_CCS": "Coal",
    "P_Coal_Lignite": "Coal", "P_Coal_Lignite_CCS": "Coal",
    "CHP_Coal_Hardcoal": "Coal", "CHP_Coal_Hardcoal_CCS": "Coal",
    "CHP_Coal_Lignite": "Coal", "CHP_Coal_Lignite_CCS": "Coal",
    # Gas
    "P_Gas_CCGT": "Gas", "P_Gas_OCGT": "Gas", "P_Gas_Engines": "Gas",
    "P_Gas_Steam": "Gas", "P_Gas_CCGT_Residual": "Gas", "P_SOFC": "Gas",
    "CHP_Gas_CCGT_Natural": "Gas",
    "CHP_Gas_CCGT_Biogas": "Gas",
    # Gas CCS
    "P_Gas_CCS": "Gas CCS", "CHP_Gas_CCGT_Natural_CCS": "Gas CCS",
    "CHP_Gas_CCGT_Biogas_CCS": "Gas CCS",
    # Other (oil, geothermal, ocean, waste, H2 turbines -> Other per user)
    "P_Oil": "Other", "CHP_Oil": "Other",
    "P_Geothermal": "Other", "P_EGS_R1": "Other", "P_EGS_R2": "Other",
    "P_EGS_R3": "Other", "P_EGS_R4": "Other",
    "P_Ocean": "Other", "CHP_WasteToEnergy": "Other",
    "P_H2_OCGT": "Other", "CHP_Hydrogen_FuelCell": "Other",
    # Nuclear
    "P_Nuclear": "Nuclear", "P_Nuclear_SMR": "Nuclear",
    # Hydro
    "P_Hydro_Reservoir": "Hydro", "P_Hydro_RoR": "Hydro",
    # Biomass
    "P_Biomass": "Biomass", "P_Biomass_CCS": "Biomass",
    "CHP_Biomass_Solid": "Biomass", "CHP_Biomass_Solid_CCS": "Biomass",
    # Solar
    "P_PV_Rooftop_Commercial": "Solar", "P_PV_Rooftop_Residential": "Solar",
    "P_PV_Utility_Avg": "Solar", "P_PV_Utility_Inf": "Solar",
    "P_PV_Utility_Opt": "Solar", "P_PV_Utility_Tracking": "Solar",
    "P_CSP": "Solar",
    # Wind
    "P_Wind_Onshore_Avg": "Wind Onshore", "P_Wind_Onshore_Inf": "Wind Onshore",
    "P_Wind_Onshore_Opt": "Wind Onshore",
    "P_Wind_Offshore_Shallow": "Wind Offshore",
    "P_Wind_Offshore_Transitional": "Wind Offshore",
    "P_Wind_Offshore_Deep": "Wind Offshore",
    # Storage
    "D_Battery_Li-Ion": "BESS",
    "D_Battery_Redox": "LDES", "D_CAES": "LDES",
    "D_PHS": "PHES",
    # dropped: X_Convert_Power, Infeasibility_Power
}

# FEL-input scheme (10 categories) -- storage excluded, H2->Hydrogen, oil->Oil
FEL_MAP = {
    "P_Coal_Hardcoal": "Coal", "P_Coal_Hardcoal_CCS": "Coal",
    "P_Coal_Lignite": "Coal", "P_Coal_Lignite_CCS": "Coal",
    "CHP_Coal_Hardcoal": "Coal", "CHP_Coal_Hardcoal_CCS": "Coal",
    "CHP_Coal_Lignite": "Coal", "CHP_Coal_Lignite_CCS": "Coal",
    "P_Gas_CCGT": "Gas", "P_Gas_OCGT": "Gas", "P_Gas_Engines": "Gas",
    "P_Gas_Steam": "Gas", "P_Gas_CCS": "Gas",
    "P_Gas_CCGT_Residual": "Gas", "P_SOFC": "Gas",
    "CHP_Gas_CCGT_Natural": "Gas", "CHP_Gas_CCGT_Biogas": "Gas",
    "CHP_Gas_CCGT_Natural_CCS": "Gas", "CHP_Gas_CCGT_Biogas_CCS": "Gas",
    "P_Hydro_Reservoir": "Hydro", "P_Hydro_RoR": "Hydro",
    "P_H2_OCGT": "Hydrogen", "CHP_Hydrogen_FuelCell": "Hydrogen",
    "P_Nuclear": "Nuclear", "P_Nuclear_SMR": "Nuclear",
    "P_Oil": "Oil", "CHP_Oil": "Oil",
    "P_Biomass": "Other Renewables", "P_Biomass_CCS": "Other Renewables",
    "CHP_Biomass_Solid": "Other Renewables",
    "CHP_Biomass_Solid_CCS": "Other Renewables",
    "P_Geothermal": "Other Renewables", "P_EGS_R1": "Other Renewables",
    "P_EGS_R2": "Other Renewables", "P_EGS_R3": "Other Renewables",
    "P_EGS_R4": "Other Renewables", "P_Ocean": "Other Renewables",
    "CHP_WasteToEnergy": "Other Renewables",
    "P_PV_Rooftop_Commercial": "Solar", "P_PV_Rooftop_Residential": "Solar",
    "P_PV_Utility_Avg": "Solar", "P_PV_Utility_Inf": "Solar",
    "P_PV_Utility_Opt": "Solar", "P_PV_Utility_Tracking": "Solar",
    "P_CSP": "Solar",
    "P_Wind_Offshore_Shallow": "Wind Offshore",
    "P_Wind_Offshore_Transitional": "Wind Offshore",
    "P_Wind_Offshore_Deep": "Wind Offshore",
    "P_Wind_Onshore_Avg": "Wind Onshore", "P_Wind_Onshore_Inf": "Wind Onshore",
    "P_Wind_Onshore_Opt": "Wind Onshore",
    # storage / convert / slack intentionally absent -> dropped
}

THINKCELL_ROWS = ["Coal", "Gas", "Gas CCS", "Other", "Nuclear", "Hydro",
                  "Biomass", "Solar", "Wind Onshore", "Wind Offshore",
                  "BESS", "LDES", "PHES"]
FEL_ROWS = ["Coal", "Gas", "Hydro", "Hydrogen", "Nuclear", "Oil",
            "Other Renewables", "Solar", "Wind Offshore", "Wind Onshore"]

# --------------------------------------------------------------------------- #
# Data access
# --------------------------------------------------------------------------- #
def distinct_scenarios(con, tables):
    """Union of distinct (Scenario, Pathway, PathwayScenario) across tables."""
    combos = set()
    for t in tables:
        try:
            for s, p, ps in con.execute(
                f"SELECT DISTINCT Scenario, Pathway, PathwayScenario FROM {t}"
            ).fetchall():
                combos.add((s, p, ps))
        except Exception:
            pass  # table absent / lacks columns
    return sorted(combos)


def choose_scenario(scenarios, label, preset=None):
    """Pick one (Scenario, Pathway, PathwayScenario) combo.

    - 0 found  -> None
    - 1 found  -> auto-selected (no prompt)
    - >1 found -> match `preset` if given, else list + prompt for stdin input
    """
    if not scenarios:
        print(f"  [warn] no scenarios found for {label}.")
        return None
    if len(scenarios) == 1:
        s, p, ps = scenarios[0]
        print(f"  {label}: single scenario -> {s} | {ps}")
        return scenarios[0]

    # non-interactive override (CLI/env): match Scenario or PathwayScenario
    if preset:
        for combo in scenarios:
            if preset == combo[0] or preset == combo[2]:
                print(f"  {label}: preset -> {combo[0]} | {combo[2]}")
                return combo
        sys.exit(f"[error] preset '{preset}' not among {label} scenarios.")

    print(f"\nAvailable {label} scenarios:")
    for i, (s, p, ps) in enumerate(scenarios, 1):
        print(f"  [{i}] Scenario={s} | Pathway={p} | PathwayScenario={ps}")
    while True:
        try:
            raw = input(
                f"Select {label} scenario [1-{len(scenarios)}] "
                f"(number or exact Scenario / PathwayScenario name): "
            ).strip()
        except EOFError:
            sys.exit(f"[error] no input for {label} scenario selection. "
                     f"Re-run interactively or pass --scenario / "
                     f"--dispatch-scenario.")
        if raw.isdigit() and 1 <= int(raw) <= len(scenarios):
            return scenarios[int(raw) - 1]
        for combo in scenarios:
            if raw == combo[0] or raw == combo[2]:
                return combo
        print("  invalid selection, try again.")


def scenario_clause(combo):
    """Build a SQL WHERE fragment from a chosen scenario combo."""
    s, p, ps = combo
    esc = lambda x: str(x).replace("'", "''")
    return (f"Scenario='{esc(s)}' AND Pathway='{esc(p)}' "
            f"AND PathwayScenario='{esc(ps)}'")


def load_results(con, where):
    """Return dict of pivoted lookups keyed by (region, category, year)."""
    cap = con.execute(
        f"SELECT Region, Technology, Type, Year, Value "
        f"FROM output_capacity WHERE {where}"
    ).fetchall()
    prod = con.execute(
        f"SELECT Region, Technology, Year, Value "
        f"FROM output_annual_production WHERE {where} AND Type='Production'"
    ).fetchall()
    demand = con.execute(
        f"SELECT Region, Year, SUM(ABS(Value)) "
        f"FROM output_annual_production WHERE {where} AND Technology='Demand' "
        f"AND Type='Use' GROUP BY Region, Year"
    ).fetchall()

    return cap, prod, {(r, y): v for r, y, v in demand}


def _aggregate(rows_tech, mapping, cols=("region", "year")):
    """rows_tech: list of (region, tech, year, value). Map tech->cat, sum."""
    out = {}  # (region, cat, year) -> value
    for region, tech, year, value in rows_tech:
        cat = mapping.get(tech)
        if cat is None:
            continue
        out[(region, cat, year)] = out.get((region, cat, year), 0.0) + value
    return out


def build_capacity_tables(cap_rows):
    """Return total / new / residual dicts keyed (region, tech, year)->GW."""
    total, new, resid = {}, {}, {}
    for region, tech, typ, year, value in cap_rows:
        if typ == "TotalCapacity":
            total[(region, tech, year)] = total.get((region, tech, year), 0) + value
        elif typ == "NewCapacity":
            new[(region, tech, year)] = new.get((region, tech, year), 0) + value
        elif typ == "ResidualCapacity":
            resid[(region, tech, year)] = resid.get((region, tech, year), 0) + value
    return total, new, resid


# --------------------------------------------------------------------------- #
# Helpers to roll up to category + region (incl. Total US)
# --------------------------------------------------------------------------- #
def cat_region_year(raw, mapping):
    """raw: dict (region,tech,year)->val. -> dict (region,cat,year)->val,
    adding synthetic 'Total US' (9 US regions) and 'Total NA' (US + Canada)."""
    agg = {}
    for (region, tech, year), val in raw.items():
        cat = mapping.get(tech)
        if cat is None:
            continue
        agg[(region, cat, year)] = agg.get((region, cat, year), 0.0) + val
    # Total US (9 US regions) and Total NA (US regions + Canada)
    extra = {}
    for (region, cat, year), val in agg.items():
        if region in US_REGIONS:
            ku = (TOTAL_US, cat, year)
            extra[ku] = extra.get(ku, 0.0) + val
        if region in US_REGIONS or region == CANADA:
            kn = (TOTAL_NA, cat, year)
            extra[kn] = extra.get(kn, 0.0) + val
    agg.update({k: agg.get(k, 0.0) + v for k, v in extra.items()})
    return agg


def prod_to_catregion(prod_rows, mapping):
    raw = {(r, t, y): v for r, t, y, v in prod_rows}
    return cat_region_year(raw, mapping)


# --------------------------------------------------------------------------- #
# File 1 & 2 : ThinkCell generation / capacities
# --------------------------------------------------------------------------- #
def fill_thinkcell(template_path, out_path, data, factor):
    """data: dict (region, cat, year) -> value (model unit). factor applied.

    Template ships 2026/2030/2040; a 2035 column is inserted so the layout
    becomes  B=2026  C=2030  D=2035  E=2040  F=helper  [G blank]  H=CAGR.
    CAGR stays 2026->2040 over 14 years (independent of the 2035 point).
    """
    from copy import copy
    wb = load_workbook(template_path)
    base = wb["Total EU"]

    # category label -> row
    cat_row = {}
    for r in range(2, base.max_row + 1):
        lab = base.cell(r, 1).value
        if isinstance(lab, str):
            cat_row[lab.strip()] = r

    # base helper formulas (col E) and which rows carry a CAGR (col G)
    helper_f = {r: base.cell(r, 5).value for r in cat_row.values()
                if base.cell(r, 5).value is not None}
    g_rows = {r for r in cat_row.values() if base.cell(r, 7).value is not None}
    n_rows = base.max_row

    COL = {2026: 2, 2030: 3, 2035: 4, 2040: 5}   # after inserting 2035 at D
    HELP_COL, CAGR_COL = 6, 8                     # F helper, H CAGR

    orig_sheets = list(wb.sheetnames)

    for disp_key in ORDER_WITH_TOTAL:
        ws = wb.copy_worksheet(base)
        ws.title = disp_name(disp_key)

        # make room: insert a blank column before the 2040 column (old D, idx4)
        ws.insert_cols(4, 1)
        ws.cell(1, 4).value = 2035
        # style the new 2035 column from the 2030 column (C)
        for r in range(1, n_rows + 1):
            src, dst = ws.cell(r, 3), ws.cell(r, 4)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)

        for cat, r in cat_row.items():
            for y, c in COL.items():
                val = data.get((disp_key, cat, y), 0.0) * factor
                ws.cell(r, c).value = round(val, 6)
            # helper (F): repoint base formula onto the 2040 column (D -> E)
            if r in helper_f:
                ws.cell(r, HELP_COL).value = helper_f[r].replace("D", "E")
            # CAGR (H): 2026 -> 2040 over 14 yrs, referencing helper F
            if r in g_rows:
                ws.cell(r, CAGR_COL).value = \
                    f'=IFERROR((F{r}/B{r})^(1/(40-26))-1,"")'

    for s in orig_sheets:
        del wb[s]
    wb.save(out_path)
    return [disp_name(k) for k in ORDER_WITH_TOTAL]
    return [disp_name(k) for k in ORDER_WITH_TOTAL]


# --------------------------------------------------------------------------- #
# File 4 : 2026 FEL input  (append region Gen+Cap sheets, bump version)
# --------------------------------------------------------------------------- #
def year_to_col(year):
    """2023->C(3) ... 2040->T(20)."""
    return 3 + (year - 2023)


def _clear_block(ws, first_row, n_tech, first_col=3, last_col=20):
    for r in range(first_row, first_row + n_tech):
        for c in range(first_col, last_col + 1):
            ws.cell(r, c).value = None


def fill_fel_input(template_path, out_path, gen_data, total, new, resid,
                   model_years):
    wb = load_workbook(template_path)
    gen_tpl = wb["Germany - Generation"]
    cap_tpl = wb["Germany - Capacity"]

    # category -> row offsets
    GEN_BASE = 9                      # rows 9..18, total 19
    CAP_BLOCKS = {                    # title row of each data block (tech row 1)
        "installed": 9, "net": 29, "retire": 49, "gross": 69,
    }
    offs = {cat: i for i, cat in enumerate(FEL_ROWS)}
    years = list(model_years)                    # annual 2025..2040

    for disp_key in ORDER_WITH_TOTAL:
        title = disp_name(disp_key)

        # ---- Generation sheet ----
        g = wb.copy_worksheet(gen_tpl)
        g.title = f"{title} - Generation"
        g["C5"] = title
        _clear_block(g, GEN_BASE, len(FEL_ROWS))
        # clear grand-total row data too (will rewrite)
        for c in range(3, 21):
            g.cell(GEN_BASE + len(FEL_ROWS), c).value = None
        for cat, off in offs.items():
            r = GEN_BASE + off
            for y in years:
                v = gen_data.get((disp_key, cat, y), 0.0) * PJ_TO_TWH
                g.cell(r, year_to_col(y)).value = round(v, 6)
        tot_r = GEN_BASE + len(FEL_ROWS)         # 19
        for y in years:
            c = year_to_col(y)
            cl = g.cell(1, c).column_letter
            g.cell(tot_r, c).value = f"=SUM({cl}{GEN_BASE}:{cl}{GEN_BASE+len(FEL_ROWS)-1})"

        # ---- Capacity sheet ----
        cp = wb.copy_worksheet(cap_tpl)
        cp.title = f"{title} - Capacity"
        for cc in (5, 25, 45, 65):
            cp.cell(cc, 3).value = title                      # C5,C25,C45,C65
        for base in CAP_BLOCKS.values():
            _clear_block(cp, base, len(FEL_ROWS))
            for c in range(3, 21):
                cp.cell(base + len(FEL_ROWS), c).value = None
        for cat, off in offs.items():
            for y in years:
                c = year_to_col(y)
                inst = total.get((disp_key, cat, y), 0.0)
                gross = new.get((disp_key, cat, y), 0.0)
                rprev = resid.get((disp_key, cat, y - 1))
                rcur = resid.get((disp_key, cat, y), 0.0)
                retire = 0.0 if rprev is None else max(0.0, rprev - rcur)
                net = gross - retire
                cp.cell(CAP_BLOCKS["installed"] + off, c).value = round(inst, 6)
                cp.cell(CAP_BLOCKS["net"] + off, c).value = round(net, 6)
                cp.cell(CAP_BLOCKS["retire"] + off, c).value = round(retire, 6)
                cp.cell(CAP_BLOCKS["gross"] + off, c).value = round(gross, 6)
        for base in CAP_BLOCKS.values():
            tot_r = base + len(FEL_ROWS)
            for y in years:
                c = year_to_col(y)
                cl = cp.cell(1, c).column_letter
                cp.cell(tot_r, c).value = \
                    f"=SUM({cl}{base}:{cl}{base+len(FEL_ROWS)-1})"

    wb.save(out_path)


# --------------------------------------------------------------------------- #
# File 3 : 260618 base case "beyond" -- NAM trade + prices sheets
# --------------------------------------------------------------------------- #
def fill_beyond(template_path, out_path, con, dcon, demand, where, disp_where):
    trade = con.execute(
        f"SELECT Region, Region2, Type, Year, Value FROM output_trade WHERE {where}"
    ).fetchall()

    # index trade
    exp = {}   # (a,b,y) -> PJ exported a->b
    imp = {}   # (a,b,y) -> PJ imported by a from b
    tcap = {}  # (a,b,y) -> GW capacity a->b
    for a, b, typ, y, v in trade:
        if typ == "Export":
            exp[(a, b, y)] = v
        elif typ == "Import":
            imp[(a, b, y)] = v
        elif typ == "Transmissions Capacity":
            tcap[(a, b, y)] = v

    regions = ORDER_NO_TOTAL          # 9 US + Canada, no Total US
    rset = set(regions)

    wb = load_workbook(template_path)

    # ---------------- NAM_Export_import (per region, MWh + %share) ---------- #
    ws = wb["NAM_Export_import"]
    ei_years = [2025, 2030, 2034, 2040]
    # clear scaffold rows 2..end
    for r in range(2, ws.max_row + 1):
        for c in range(1, 6):
            ws.cell(r, c).value = None
    row = 2
    for reg in regions:
        disp = DISPLAY[reg]
        for y in ei_years:
            gexp = sum(v for (a, b, yr), v in exp.items() if a == reg and yr == y and b in rset)
            gimp = sum(v for (a, b, yr), v in imp.items() if a == reg and yr == y and b in rset)
            net = gexp - gimp
            cons = demand.get((reg, y))
            for ind, val_pj in [("Gross export", gexp), ("Gross import", gimp),
                                ("Net export", net)]:
                ws.cell(row, 1).value = ind
                ws.cell(row, 2).value = disp
                ws.cell(row, 3).value = "MWh"
                ws.cell(row, 4).value = y
                ws.cell(row, 5).value = round(val_pj * PJ_TO_MWH, 3)
                row += 1
            for ind, val_pj in [("Gross export", gexp), ("Gross import", gimp),
                                ("Net export", net)]:
                ws.cell(row, 1).value = ind
                ws.cell(row, 2).value = disp
                ws.cell(row, 3).value = "% Share of total domestic consumption"
                ws.cell(row, 4).value = y
                ws.cell(row, 5).value = (round(val_pj / cons, 6)
                                         if cons not in (None, 0) else None)
                row += 1

    # ---------------- NAM_Pairwise trade (unordered pairs, TWh) ------------- #
    ws = wb["NAM_Pairwise trade"]
    for r in range(2, ws.max_row + 1):
        for c in range(1, 7):
            ws.cell(r, c).value = None
    pw_years = [2025, 2030, 2034, 2040]
    row = 2
    for i in range(len(regions)):
        for j in range(i + 1, len(regions)):
            a, b = regions[i], regions[j]
            da, db = DISPLAY[a], DISPLAY[b]
            pair = f"{da}-{db}"
            # only emit a pair that has any trade or capacity
            has = any((a, b, y) in exp or (b, a, y) in exp or (a, b, y) in tcap
                      for y in pw_years)
            if not has:
                continue
            for y in pw_years:
                gexp = exp.get((a, b, y), 0.0)            # a -> b
                gimp = imp.get((a, b, y), exp.get((b, a, y), 0.0))  # into a from b
                net = gexp - gimp
                for ind, val, note in [
                    ("Gross export", gexp, f"Export from {da} to {db}"),
                    ("Gross import", gimp, f"Import from {db} to {da}"),
                    ("Net export", net, f"Net export {da} to {db}"),
                ]:
                    ws.cell(row, 1).value = pair
                    ws.cell(row, 2).value = ind
                    ws.cell(row, 3).value = "TWh"
                    ws.cell(row, 4).value = y
                    ws.cell(row, 5).value = round(val * PJ_TO_TWH, 6)
                    ws.cell(row, 6).value = note
                    row += 1

    # ---------------- NAM_IC Capacity (unordered pairs, GW) ----------------- #
    ws = wb["NAM_IC Capacity"]
    for r in range(2, ws.max_row + 1):
        for c in range(1, 7):
            ws.cell(r, c).value = None
    # every model year (annual), not stepped
    ic_years = sorted({y for (_, _, y) in tcap})
    row = 2
    for i in range(len(regions)):
        for j in range(i + 1, len(regions)):
            a, b = regions[i], regions[j]
            da, db = DISPLAY[a], DISPLAY[b]
            has = any((a, b, y) in tcap or (b, a, y) in tcap for y in ic_years)
            if not has:
                continue
            pair = f"{da}-{db}"
            for y in ic_years:
                v = tcap.get((a, b, y), tcap.get((b, a, y)))
                if v is None:
                    continue
                ws.cell(row, 1).value = "Interconnector capacity"
                ws.cell(row, 2).value = pair
                ws.cell(row, 3).value = y
                ws.cell(row, 4).value = "GW"
                ws.cell(row, 5).value = round(v, 4)
                ws.cell(row, 6).value = f"Capacity from {da} to {db}"
                row += 1

    # ---------------- NAM_Prices (per region, EUR/MWh) ---------------------- #
    ws = wb["NAM_Prices"]
    for r in range(2, ws.max_row + 1):
        for c in range(1, 6):
            ws.cell(r, c).value = None
    price_years = [2025, 2030, 2032, 2034, 2036, 2038, 2040]
    dcols = [c[0] for c in dcon.execute("DESCRIBE dispatch_summary").fetchall()]
    has_region = "Region" in dcols
    dfilter = f" WHERE {disp_where}" if disp_where else ""
    if has_region:
        prows = dcon.execute(
            f"SELECT Region, Year, AvgPrice_EURMWh FROM dispatch_summary{dfilter}"
        ).fetchall()
        price = {(r, y): p for r, y, p in prows}
    else:
        prows = dcon.execute(
            f"SELECT Year, AvgPrice_EURMWh FROM dispatch_summary{dfilter}"
        ).fetchall()
        nat = {y: p for y, p in prows}
        price = {(reg, y): nat.get(y) for reg in regions for y in nat}
        print("  [warn] dispatch_summary has no Region column -> "
              "national AvgPrice applied to every region.")
    row = 2
    for reg in regions:
        disp = DISPLAY[reg]
        for y in price_years:
            p = price.get((reg, y))
            if p is None:
                continue
            ws.cell(row, 1).value = "Wholesale power prices"
            ws.cell(row, 2).value = disp
            ws.cell(row, 3).value = y
            ws.cell(row, 4).value = "EUR/MWh"
            ws.cell(row, 5).value = round(p, 4)
            row += 1

    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def audit_technologies(con):
    """Print DB technologies not covered by either mapping table."""
    # non-generation rows handled elsewhere / deliberately dropped
    IGNORE = {"X_Convert_Power", "Infeasibility_Power", "Demand", "Trade"}
    techs = set()
    for t in ("output_capacity", "output_annual_production"):
        try:
            techs |= {r[0] for r in
                      con.execute(f"SELECT DISTINCT Technology FROM {t}").fetchall()}
        except Exception:
            pass
    known = set(THINKCELL_MAP) | set(FEL_MAP)
    new = sorted(techs - known - IGNORE)
    # informational: mapped in ThinkCell but not FEL (usually intentional,
    # e.g. storage is excluded from the FEL input file on purpose)
    tc_only = sorted((set(THINKCELL_MAP) & techs) - set(FEL_MAP) - IGNORE)

    print(f"\nTechnologies in DB: {len(techs)}  "
          f"(ignored: {sorted(IGNORE)})\n")
    print(f"NEW / UNMAPPED - not in either map ({len(new)}):")
    for t in new:
        print(f"    {t}")
    if not new:
        print("    (none - every technology is mapped)")
    if tc_only:
        print(f"\n[info] mapped in ThinkCell but not FEL ({len(tc_only)}) "
              f"- normally intentional (storage excluded from FEL):")
        for t in tc_only:
            print(f"    {t}")
    if new:
        print("\nPaste the NEW list back to me and I'll add each to both "
              "mapping tables (I'll ask about any ambiguous ones).")
    con.close()


def main():
    import argparse
    ap = argparse.ArgumentParser(
        description="Convert GENeSYS-MOD NA results into FEL templates.")
    ap.add_argument("--scenario", default=os.environ.get("GMOD_SCENARIO"),
                    help="Results-DB scenario (exact Scenario or "
                         "PathwayScenario name) to skip the prompt.")
    ap.add_argument("--dispatch-scenario",
                    default=os.environ.get("GMOD_DISPATCH_SCENARIO"),
                    help="Dispatch-DB scenario to skip the prompt.")
    ap.add_argument("--audit-tech", action="store_true",
                    help="List DB technologies not covered by the mapping "
                         "tables (then exit). Use to find newly added techs.")
    args = ap.parse_args()

    if args.audit_tech:
        if not os.path.exists(RESULTS_DB):
            sys.exit(f"Missing database: {RESULTS_DB}")
        audit_technologies(duckdb.connect(RESULTS_DB, read_only=True))
        return

    for p in (RESULTS_DB, DISPATCH_DB):
        if not os.path.exists(p):
            sys.exit(f"Missing database: {p}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    con = duckdb.connect(RESULTS_DB, read_only=True)
    dcon = duckdb.connect(DISPATCH_DB, read_only=True)

    # ----- scenario selection (one choice applied to ALL results tables) -----
    res_combo = choose_scenario(
        distinct_scenarios(con, ["output_capacity",
                                 "output_annual_production", "output_trade"]),
        "results", preset=args.scenario)
    if res_combo is None:
        sys.exit("[error] no scenarios in results DB.")
    res_where = scenario_clause(res_combo)

    disp_combo = choose_scenario(
        distinct_scenarios(dcon, ["dispatch_summary"]),
        "dispatch", preset=args.dispatch_scenario)
    disp_where = scenario_clause(disp_combo) if disp_combo else None

    print(f"\n  results scenario : {res_combo[0]} | {res_combo[2]}")
    print(f"  dispatch scenario: "
          f"{disp_combo[0] if disp_combo else '(none)'}\n")

    cap_rows, prod_rows, demand = load_results(con, res_where)
    total_raw, new_raw, resid_raw = build_capacity_tables(cap_rows)

    model_years = sorted({y for (_, _, y) in total_raw})

    # ThinkCell aggregations (incl Total US)
    gen_tc = prod_to_catregion(prod_rows, THINKCELL_MAP)
    cap_tc = cat_region_year(total_raw, THINKCELL_MAP)

    # FEL aggregations (incl Total US)
    gen_fel = prod_to_catregion(prod_rows, FEL_MAP)
    total_fel = cat_region_year(total_raw, FEL_MAP)
    new_fel = cat_region_year(new_raw, FEL_MAP)
    resid_fel = cat_region_year(resid_raw, FEL_MAP)

    # ----- File 1: Generation ThinkCell (PJ -> GWh) -----
    out1 = os.path.join(OUTPUT_DIR, "FEL_BaseCase_NA_Generation_ThinkCell.xlsx")
    fill_thinkcell(TPL_GEN, out1, gen_tc, PJ_TO_GWH)
    print(f"  wrote {os.path.basename(out1)}")

    # ----- File 2: Capacities ThinkCell (GW -> MW) -----
    out2 = os.path.join(OUTPUT_DIR, "FEL_BaseCase_NA_Capacities_ThinkCell.xlsx")
    fill_thinkcell(TPL_CAP, out2, cap_tc, GW_TO_MW)
    print(f"  wrote {os.path.basename(out2)}")

    # ----- File 4: FEL input v03 (Gen TWh, Cap GW) -----
    today = datetime.date.today().strftime("%y%m%d")
    out4 = os.path.join(
        OUTPUT_DIR,
        f"2026_FEL_ElectricityGenerationByTechnology_Input_v03-{today}.xlsx")
    fill_fel_input(TPL_FEL_INPUT, out4, gen_fel, total_fel, new_fel, resid_fel,
                   model_years)
    print(f"  wrote {os.path.basename(out4)}")

    # ----- File 3: beyond (NAM trade + prices) -----
    out3 = os.path.join(
        OUTPUT_DIR,
        "260618_Base_case_results_beyond_Capacity_and_generation_v3.xlsx")
    fill_beyond(TPL_BEYOND, out3, con, dcon, demand, res_where, disp_where)
    print(f"  wrote {os.path.basename(out3)}")

    con.close()
    dcon.close()
    print("Done.")


if __name__ == "__main__":
    main()
