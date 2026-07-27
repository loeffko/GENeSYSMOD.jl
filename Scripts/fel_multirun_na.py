#!/usr/bin/env python3
"""
fel_multirun_na.py

Multi-scenario version of the 2026_FEL_ElectricityGenerationByTechnology
output: one Generation + one Capacity sheet per region PER SELECTED RUN,
NA regions only (all EU template sheets are dropped).

Output file: Output/FEL_AllRuns_NA_GT_Results.xlsx
Sheet names:  <run>|<Region>|Gen   and   <run>|<Region>|Cap
              (31-char Excel limit -> run label truncated if needed; the full
              label is written into the sheet header cells C5/C25/...)

Scenario mapping CSV (repeat runs without number lookup):
    Scripts/scenario_mapping.csv with columns:
        investment run, dispatch run, name
    - 'investment run' matches Scenario or PathwayScenario in the results DB
    - 'dispatch run' is carried for reuse by other scripts (ignored here)
    - 'name' is the display label used in sheet names / headers
    At start the script asks whether to read the mapping from this file; on an
    interactive selection it offers to save the mapping for next time.

Run:
    python fel_multirun_na.py
    python fel_multirun_na.py --use-mapping            # read CSV, no prompt
    python fel_multirun_na.py --scenarios all
    python fel_multirun_na.py --scenarios "244_invLimit_ramping_v01" \
        --names "244_invLimit_ramping_v01=Base"
"""

import os
import sys
import csv
import argparse
import duckdb
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(SCRIPT_DIR, "templates")
RESULTS_DIR = os.path.join(SCRIPT_DIR, "..", "Results")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "Output")
RESULTS_DB = os.path.join(RESULTS_DIR, "genesysmod_db.duckdb")
MAPPING_CSV = os.path.join(SCRIPT_DIR, "scenario_mapping.csv")
TPL_FEL = os.path.join(TEMPLATE_DIR,
                       "2026_FEL_ElectricityGenerationByTechnology_Input_v02-260619.xlsx")
OUT_NAME = "FEL_AllRuns_NA_GT_Results.xlsx"

# --------------------------------------------------------------------------- #
# Regions
# --------------------------------------------------------------------------- #
US_REGIONS = ["California", "ERCOT", "MISO", "NewEngland", "NewYork",
              "PJM", "SERC", "SPP", "WECC"]
CANADA = "Canada"
TOTAL_US = "Total US"
TOTAL_NA = "Total NA"
DISPLAY = {"California": "California", "ERCOT": "ERCOT", "MISO": "MISO",
           "NewEngland": "New England", "NewYork": "New York", "PJM": "PJM",
           "SERC": "SERC", "SPP": "SPP", "WECC": "WECC", "Canada": "Canada"}
ORDER = [TOTAL_NA, TOTAL_US] + US_REGIONS + [CANADA]


def disp_name(key):
    return DISPLAY.get(key, key)


# --------------------------------------------------------------------------- #
# FEL technology mapping (10 categories, storage excluded)
# --------------------------------------------------------------------------- #
FEL_MAP = {
    "P_Coal_Hardcoal": "Coal", "P_Coal_Hardcoal_CCS": "Coal",
    "P_Coal_Lignite": "Coal", "P_Coal_Lignite_CCS": "Coal",
    "CHP_Coal_Hardcoal": "Coal", "CHP_Coal_Hardcoal_CCS": "Coal",
    "CHP_Coal_Lignite": "Coal", "CHP_Coal_Lignite_CCS": "Coal",
    "P_Gas_CCGT": "Gas", "P_Gas_OCGT": "Gas", "P_Gas_Engines": "Gas",
    "P_Gas_Steam": "Gas", "P_Gas_CCS": "Gas",
    "P_Gas_CCGT_Residual": "Gas", "P_SOFC": "Gas",
    "CHP_Gas_CCGT_Natural": "Gas", "CHP_Gas_CCGT_Biogas": "Gas",
    "CHP_Gas_CCGT_Natural_CCS": "Gas", "CHP_Gas_CCGT_Biogas_CCS": "Gas",
    "P_Hydro_Reservoir": "Hydro", "P_Hydro_RoR": "Hydro",
    "P_H2_OCGT": "Hydrogen", "CHP_Hydrogen_FuelCell": "Hydrogen",
    "P_Nuclear": "Nuclear", "P_Nuclear_SMR": "Nuclear",
    "P_Oil": "Oil", "CHP_Oil": "Oil",
    "P_Biomass": "Other Renewables", "P_Biomass_CCS": "Other Renewables",
    "CHP_Biomass_Solid": "Other Renewables",
    "CHP_Biomass_Solid_CCS": "Other Renewables",
    "P_Geothermal": "Other Renewables", "P_EGS_R1": "Other Renewables",
    "P_EGS_R2": "Other Renewables", "P_EGS_R3": "Other Renewables",
    "P_EGS_R4": "Other Renewables", "P_Ocean": "Other Renewables",
    "CHP_WasteToEnergy": "Other Renewables",
    "P_PV_Rooftop_Commercial": "Solar", "P_PV_Rooftop_Residential": "Solar",
    "P_PV_Utility_Avg": "Solar", "P_PV_Utility_Inf": "Solar",
    "P_PV_Utility_Opt": "Solar", "P_PV_Utility_Tracking": "Solar",
    "P_CSP": "Solar",
    "P_Wind_Offshore_Shallow": "Wind Offshore",
    "P_Wind_Offshore_Transitional": "Wind Offshore",
    "P_Wind_Offshore_Deep": "Wind Offshore",
    "P_Wind_Onshore_Avg": "Wind Onshore", "P_Wind_Onshore_Inf": "Wind Onshore",
    "P_Wind_Onshore_Opt": "Wind Onshore",
}
FEL_ROWS = ["Coal", "Gas", "Hydro", "Hydrogen", "Nuclear", "Oil",
            "Other Renewables", "Solar", "Wind Offshore", "Wind Onshore"]

PJ_TO_TWH = 1.0 / 3.6

# --------------------------------------------------------------------------- #
# Scenario selection + mapping CSV
# --------------------------------------------------------------------------- #
def distinct_scenarios(con, tables):
    combos = set()
    for t in tables:
        try:
            for s, p, ps in con.execute(
                f"SELECT DISTINCT Scenario, Pathway, PathwayScenario FROM {t}"
            ).fetchall():
                combos.add((s, p, ps))
        except Exception:
            pass
    return sorted(combos)


def scenario_clause(combo):
    s, p, ps = combo
    esc = lambda x: str(x).replace("'", "''")
    return (f"Scenario='{esc(s)}' AND Pathway='{esc(p)}' "
            f"AND PathwayScenario='{esc(ps)}'")


def _match(scenarios, token):
    token = token.strip()
    if token.isdigit() and 1 <= int(token) <= len(scenarios):
        return scenarios[int(token) - 1]
    for c in scenarios:
        if token == c[0] or token == c[2]:
            return c
    return None


def read_mapping_csv(path):
    """Return list of (investment_run, dispatch_run, name)."""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        for row in reader:
            if not row or not row[0].strip():
                continue
            inv = row[0].strip()
            dis = row[1].strip() if len(row) > 1 else ""
            name = row[2].strip() if len(row) > 2 and row[2].strip() else inv
            rows.append((inv, dis, name))
    return rows


def write_mapping_csv(path, entries):
    with open(path, "w", newline="", encoding="utf-8") as f:
        wcsv = csv.writer(f)
        wcsv.writerow(["investment run", "dispatch run", "name"])
        for inv, dis, name in entries:
            wcsv.writerow([inv, dis, name])


def select_runs(scenarios, args):
    """Return list of (combo, label). Sources, in priority:
    1. --scenarios/--names CLI
    2. mapping CSV (if --use-mapping, or user says yes at the prompt)
    3. interactive multi-select + name prompts (offering to save the CSV)
    """
    # --- CLI preset ---
    if args.scenarios:
        if args.scenarios.strip().lower() == "all":
            chosen = list(scenarios)
        else:
            chosen = [_match(scenarios, t) for t in args.scenarios.split(",")]
            if any(c is None for c in chosen):
                sys.exit(f"[error] unknown scenario in --scenarios.")
        names = {}
        if args.names:
            for part in args.names.split(";"):
                if "=" in part:
                    k, v = part.split("=", 1)
                    names[k.strip()] = v.strip()
        return [(c, names.get(c[0], names.get(c[2], c[0]))) for c in chosen]

    # --- mapping CSV ---
    use_csv = args.use_mapping
    if not use_csv and os.path.exists(MAPPING_CSV) and sys.stdin.isatty():
        ans = input(f"Read run selection from "
                    f"{os.path.basename(MAPPING_CSV)}? [y/N]: ").strip().lower()
        use_csv = ans == "y"
    if use_csv:
        if not os.path.exists(MAPPING_CSV):
            sys.exit(f"[error] mapping file not found: {MAPPING_CSV}")
        entries = read_mapping_csv(MAPPING_CSV)
        out = []
        for inv, _dis, name in entries:
            c = _match(scenarios, inv)
            if c is None:
                sys.exit(f"[error] mapping entry '{inv}' not in results DB.")
            out.append((c, name))
        if not out:
            sys.exit("[error] mapping file has no entries.")
        print(f"  {len(out)} run(s) read from mapping file.")
        return out

    # --- interactive ---
    if len(scenarios) == 1:
        c = scenarios[0]
        print(f"  single scenario -> {c[0]}")
        label = c[0]
        if sys.stdin.isatty():
            raw = input(f"Display name for '{c[0]}' [default {c[0]}]: ").strip()
            label = raw or c[0]
        chosen = [(c, label)]
    else:
        print("\nAvailable scenarios:")
        for i, (s, p, ps) in enumerate(scenarios, 1):
            print(f"  [{i}] Scenario={s} | PathwayScenario={ps}")
        while True:
            try:
                raw = input("Select runs (comma-separated numbers/names, or "
                            "'all'): ").strip()
            except EOFError:
                sys.exit("[error] no input. Pass --scenarios or --use-mapping.")
            if raw.lower() == "all":
                picked = list(scenarios)
            else:
                picked = [_match(scenarios, t) for t in raw.split(",")
                          if t.strip()]
                if not picked or any(c is None for c in picked):
                    print("  invalid selection, try again.")
                    continue
            seen, uniq = set(), []
            for c in picked:
                if c not in seen:
                    seen.add(c); uniq.append(c)
            break
        chosen = []
        for c in uniq:
            raw = input(f"Display name for '{c[0]}' [default {c[0]}]: ").strip()
            chosen.append((c, raw or c[0]))

    # offer to save mapping
    if sys.stdin.isatty():
        ans = input(f"Save this selection to "
                    f"{os.path.basename(MAPPING_CSV)} for next time? "
                    f"[y/N]: ").strip().lower()
        if ans == "y":
            write_mapping_csv(MAPPING_CSV,
                              [(c[0], "", label) for c, label in chosen])
            print(f"  mapping saved -> {MAPPING_CSV}")
    return chosen


# --------------------------------------------------------------------------- #
# Data loading / aggregation (per scenario)
# --------------------------------------------------------------------------- #
def load_scenario(con, where):
    cap = con.execute(
        f"SELECT Region, Technology, Type, Year, Value FROM output_capacity "
        f"WHERE {where}").fetchall()
    prod = con.execute(
        f"SELECT Region, Technology, Year, Value FROM output_annual_production "
        f"WHERE {where} AND Type='Production'").fetchall()

    total, new, resid = {}, {}, {}
    for region, tech, typ, year, value in cap:
        d = {"TotalCapacity": total, "NewCapacity": new,
             "ResidualCapacity": resid}.get(typ)
        if d is not None:
            d[(region, tech, year)] = d.get((region, tech, year), 0) + value
    prod_raw = {(r, t, y): v for r, t, y, v in prod}
    return total, new, resid, prod_raw


def cat_region_year(raw, mapping):
    agg = {}
    for (region, tech, year), val in raw.items():
        cat = mapping.get(tech)
        if cat is None:
            continue
        agg[(region, cat, year)] = agg.get((region, cat, year), 0.0) + val
    extra = {}
    for (region, cat, year), val in agg.items():
        if region in US_REGIONS:
            ku = (TOTAL_US, cat, year)
            extra[ku] = extra.get(ku, 0.0) + val
        if region in US_REGIONS or region == CANADA:
            kn = (TOTAL_NA, cat, year)
            extra[kn] = extra.get(kn, 0.0) + val
    agg.update({k: agg.get(k, 0.0) + v for k, v in extra.items()})
    return agg


# --------------------------------------------------------------------------- #
# Sheet writing: one Gen + one Cap sheet per region, one table block per run
# stacked downwards, labelled via the block's Scenario cell (C4 etc.).
# --------------------------------------------------------------------------- #
from copy import copy as _copy

GEN_BASE = 9                 # first tech row in the template's generation block
CAP_BLOCKS = {"installed": 9, "net": 29, "retire": 49, "gross": 69}
GEN_TPL_ROWS = (2, 19)       # replicated row span, generation block
CAP_TPL_ROWS = (2, 79)       # replicated row span, capacity 4-block group
GEN_PITCH = 20               # row offset per additional run (gen)
CAP_PITCH = 80               # row offset per additional run (cap)
META_SCEN_ROWS_GEN = [4]     # rows whose C cell carries the Scenario label
META_CNTY_ROWS_GEN = [5]
META_SCEN_ROWS_CAP = [4, 24, 44, 64]
META_CNTY_ROWS_CAP = [5, 25, 45, 65]
FIRST_COL, LAST_COL = 2, 20  # B..T


def year_to_col(year):
    return 3 + (year - 2023)          # 2023 -> C


def _replicate_block(ws, tpl, row_span, offset):
    """Copy styles for the whole block and values for labels/headers/meta from
    the template sheet into ws at +offset rows. Data cells stay empty (they
    are written fresh); template data values and formulas are NOT copied."""
    r0, r1 = row_span
    for r in range(r0, r1 + 1):
        for c in range(FIRST_COL, LAST_COL + 1):
            src = tpl.cell(r, c)
            dst = ws.cell(r + offset, c)
            if src.has_style:
                dst.font = _copy(src.font)
                dst.fill = _copy(src.fill)
                dst.border = _copy(src.border)
                dst.alignment = _copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = _copy(src.protection)
            # values: labels (col B), year header rows, meta/pivot label rows
            if c == 2 or (isinstance(src.value, (int, str)) and r in
                          (7, 8, 27, 28, 47, 48, 67, 68) ):
                dst.value = src.value


def _clear_data(ws, first_row, n_tech):
    for r in range(first_row, first_row + n_tech + 1):   # + total row
        for c in range(3, LAST_COL + 1):
            ws.cell(r, c).value = None


def _write_table(ws, base_row, values_by_cat, model_years, rounder=6):
    offs = {cat: i for i, cat in enumerate(FEL_ROWS)}
    for cat, off in offs.items():
        r = base_row + off
        for y in model_years:
            ws.cell(r, year_to_col(y)).value = round(values_by_cat.get((cat, y), 0.0), rounder)
    tot_r = base_row + len(FEL_ROWS)
    for y in model_years:
        c = year_to_col(y)
        cl = ws.cell(1, c).column_letter
        ws.cell(tot_r, c).value = \
            f"=SUM({cl}{base_row}:{cl}{base_row + len(FEL_ROWS) - 1})"


def build_region_sheets(wb, gen_tpl, cap_tpl, runs_data):
    """runs_data: list of (label, gen_fel, total_fel, new_fel, resid_fel,
    model_years) aggregated dicts keyed (region, cat, year)."""
    for key in ORDER:
        rd = disp_name(key)

        # ---------- Generation sheet ----------
        g = wb.copy_worksheet(gen_tpl)
        g.title = f"{rd} - Generation"
        for k, run in enumerate(runs_data):
            label, gen_fel, _t, _n, _r, model_years = run
            off = k * GEN_PITCH
            if k > 0:
                _replicate_block(g, gen_tpl, GEN_TPL_ROWS, off)
            g.cell(4 + off, 3).value = label          # C4: Scenario
            g.cell(5 + off, 3).value = rd             # C5: Country/Region
            _clear_data(g, GEN_BASE + off, len(FEL_ROWS))
            vals = {(cat, y): v * PJ_TO_TWH
                    for (reg, cat, y), v in gen_fel.items() if reg == key}
            _write_table(g, GEN_BASE + off, vals, model_years)

        # ---------- Capacity sheet ----------
        cp = wb.copy_worksheet(cap_tpl)
        cp.title = f"{rd} - Capacity"
        for k, run in enumerate(runs_data):
            label, _gen, total_fel, new_fel, resid_fel, model_years = run
            off = k * CAP_PITCH
            if k > 0:
                _replicate_block(cp, cap_tpl, CAP_TPL_ROWS, off)
            for mr in META_SCEN_ROWS_CAP:
                cp.cell(mr + off, 3).value = label
            for mr in META_CNTY_ROWS_CAP:
                cp.cell(mr + off, 3).value = rd
            # derive the four metric tables for this region
            inst, net, retire, gross = {}, {}, {}, {}
            cats = {cat for (reg, cat, y) in total_fel if reg == key}
            years = set(model_years)
            for cat in cats | {c for (reg, c, y) in new_fel if reg == key}:
                for y in years:
                    i = total_fel.get((key, cat, y), 0.0)
                    gr = new_fel.get((key, cat, y), 0.0)
                    rprev = resid_fel.get((key, cat, y - 1))
                    rcur = resid_fel.get((key, cat, y), 0.0)
                    rt = 0.0 if rprev is None else max(0.0, rprev - rcur)
                    inst[(cat, y)] = i
                    gross[(cat, y)] = gr
                    retire[(cat, y)] = rt
                    net[(cat, y)] = gr - rt
            for name, table in [("installed", inst), ("net", net),
                                ("retire", retire), ("gross", gross)]:
                base = CAP_BLOCKS[name] + off
                _clear_data(cp, base, len(FEL_ROWS))
                _write_table(cp, base, table, model_years)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(
        description="Multi-run FEL Gen/Cap output (NA only, EU sheets dropped).")
    ap.add_argument("--scenarios", default=None,
                    help="'all' or comma-separated Scenario/PathwayScenario names.")
    ap.add_argument("--names", default=None,
                    help="Optional 'scen=Label;...' display names (with --scenarios).")
    ap.add_argument("--use-mapping", action="store_true",
                    help=f"Read run selection from {os.path.basename(MAPPING_CSV)} "
                         f"without asking.")
    ap.add_argument("-o", "--output", default=None)
    args = ap.parse_args()

    for p in (RESULTS_DB, TPL_FEL):
        if not os.path.exists(p):
            sys.exit(f"Missing file: {p}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    con = duckdb.connect(RESULTS_DB, read_only=True)
    scenarios = distinct_scenarios(
        con, ["output_capacity", "output_annual_production"])
    runs = select_runs(scenarios, args)

    print(f"\n  runs selected: {len(runs)}")
    for c, label in runs:
        print(f"    - {c[0]}  ->  '{label}'")

    wb = load_workbook(TPL_FEL)
    gen_tpl = wb["Germany - Generation"]
    cap_tpl = wb["Germany - Capacity"]
    template_sheets = list(wb.sheetnames)          # all EU sheets: drop later

    runs_data = []
    for combo, label in runs:
        where = scenario_clause(combo)
        total, new, resid, prod_raw = load_scenario(con, where)
        model_years = sorted({y for (_, _, y) in total})
        runs_data.append((label,
                          cat_region_year(prod_raw, FEL_MAP),
                          cat_region_year(total, FEL_MAP),
                          cat_region_year(new, FEL_MAP),
                          cat_region_year(resid, FEL_MAP),
                          model_years))
    con.close()

    build_region_sheets(wb, gen_tpl, cap_tpl, runs_data)

    # drop every original (EU) template sheet -> NA only
    for s in template_sheets:
        del wb[s]

    out_path = args.output or os.path.join(OUTPUT_DIR, OUT_NAME)
    wb.save(out_path)
    n_sheets = len(ORDER) * 2
    print(f"\n  wrote {os.path.basename(out_path)} "
          f"({n_sheets} sheets, {len(runs)} run block(s) each)")
    print("Done.")


if __name__ == "__main__":
    main()
